#!/usr/bin/env python3
"""Validate that a dataset can actually support a requested representation.

Reads a CSV (or one sheet of an XLSX when openpyxl is available) and checks the
requested representation against the minimum requirements in
``references/data-task-fit.md``.

Column roles are declared, not guessed. Type inference alone cannot tell a
measurement from an identifier — a table of student IDs and course codes is two
numeric columns, and inference would certify it for a scatter plot. So the
representations whose meaning depends on which column plays which part require
``--x``/``--y``/``--order`` and friends, and the validator then checks that the
named column actually holds the kind of data the role needs.

This is a structural screen. It confirms that named columns exist and that
types, roles, pairing, levels, and observation counts support the
representation. It cannot judge whether the resulting chart answers the
student's question, and it cannot confirm semantic conditions it has no way to
see — that pie slices are parts of the intended whole, or that a bin width
suits the distribution. Those stay with a qualified human.

Exit codes:
  0: the dataset supports the declared representation
  1: file, parsing, or structural error
  2: data-task fit gaps detected
"""

from __future__ import annotations

import argparse
import csv
import math
import re
import sys
from pathlib import Path
from typing import NamedTuple

from _tabular import emit_report, normalize, parse_iso_date

QUANT = "quantitative"
CATEG = "categorical"
ORDER = "ordered"

EMPTY_ROLES: dict[str, str] = {}


class Requirement(NamedTuple):
    """What a representation needs, mirroring references/data-task-fit.md.

    `roles` maps a role name to the column kind it requires. `require_roles`
    marks representations where inference is unsafe, so the roles must be
    declared explicitly rather than guessed from column types. `min_levels`,
    `nonnegative`, and `unique` carry the conditions that make a representation
    meaningful rather than merely constructible: a comparison needs more than
    one group, a pie needs nonnegative parts that each appear once.
    """

    roles: dict[str, str]
    min_rows: int
    optional_roles: dict[str, str] = EMPTY_ROLES
    require_roles: bool = False
    paired: bool = False
    min_levels: int = 0
    nonnegative: tuple[str, ...] = ()
    unique: tuple[str, ...] = ()


REPRESENTATIONS: dict[str, Requirement] = {
    "scatter": Requirement(
        {"x": QUANT, "y": QUANT}, 3, require_roles=True, paired=True
    ),
    "correlation": Requirement(
        {"x": QUANT, "y": QUANT}, 3, require_roles=True, paired=True
    ),
    "regression": Requirement(
        {"x": QUANT, "y": QUANT}, 3, require_roles=True, paired=True
    ),
    "line": Requirement(
        {"order": ORDER, "y": QUANT}, 3, require_roles=True, paired=True
    ),
    "bar": Requirement({"category": CATEG, "value": QUANT}, 2),
    "pie": Requirement(
        {"category": CATEG, "value": QUANT},
        2,
        min_levels=2,
        nonnegative=("value",),
        unique=("category",),
    ),
    # Two binned or categorical dimensions, one value per cell. One dimension
    # plus a value is a bar chart, so the second dimension must be named.
    "heatmap": Requirement(
        {"category": CATEG, "series": CATEG, "value": QUANT},
        4,
        require_roles=True,
        min_levels=2,
    ),
    "grouped-comparison": Requirement(
        {"category": CATEG, "value": QUANT}, 4, require_roles=True, min_levels=2
    ),
    "box": Requirement({"value": QUANT}, 5, optional_roles={"category": CATEG}),
    "histogram": Requirement({"value": QUANT}, 5),
    "mean": Requirement({"value": QUANT}, 2),
    "standard-deviation": Requirement({"value": QUANT}, 3),
    "uncertainty": Requirement({"value": QUANT}, 3),
}

ROLE_FLAGS = ("x", "y", "category", "series", "value", "order")

MISSING_TOKENS = {"", "na", "n/a", "nan", "null", "none", "-", "--", "."}
# Headers that name a key rather than a measurement.
IDENTIFIER_HEADER = re.compile(
    r"(^|[\s_-])(id|ids|code|codes|key|keys|number|no|num|index|uuid|roster)$",
    re.IGNORECASE,
)
# Headers that name a measured quantity. A unit suffix or a measure word is
# strong evidence against the identifier heuristic below: dose_mg = 0, 1, 2, 3
# and time_s = 1, 2, 3, 4 are consecutive small integers and also real data.
MEASUREMENT_HEADER = re.compile(
    # Unit suffixes, all at least two characters. Single-letter units are
    # deliberately absent: section_a and participant_n are not amperes and
    # newtons, and admitting them let a 1..n key pass as a measurement. Columns
    # that really do carry a one-letter unit — current_a, force_n, temp_c,
    # time_s — are already covered by the measure-word stem below.
    r"(_(?:mg|kg|mm|cm|km|nm|um|ft|lb|oz|ms|us|ns|sec|min|hr|ml|dl|mol|mmol|"
    r"ppm|ppb|pct|deg|rad|hz|khz|mhz|pa|kpa|mpa|psi|kwh|m2|m3|cm2|cm3)$"
    # Measure words are matched at underscore- and hyphen-delimited boundaries
    # as well as word boundaries, so trial_count and elapsed_time read as the
    # measurements they are.
    r"|(?:^|[\s_-])(?:dose|time|temp|temperature|mass|weight|length|height|"
    r"width|depth|volume|conc|concentration|pressure|force|speed|velocity|"
    r"current|voltage|power|energy|freq|frequency|angle|distance|dist|age|"
    r"score|count|rate|yield|density|ph|absorbance|intensity|elapsed|duration|"
    r"reading|measurement)(?:$|[\s_-]))",
    re.IGNORECASE,
)
ORDERED_HEADER = re.compile(
    r"(date|time|timestamp|year|month|week|day|hour|minute|second|step|trial|"
    r"session|order|sequence|period|epoch|cycle|iteration|run)",
    re.IGNORECASE,
)


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description=(
            "Check that a dataset supports a requested chart or statistic before "
            "the activity is released to students."
        ),
        epilog=(
            "Examples:\n"
            "  validate_dataset.py data.csv --representation scatter "
            "--x mass_kg --y extension_mm\n"
            "  validate_dataset.py data.csv --representation bar "
            "--category region --value total\n"
            "  validate_dataset.py data.csv --representation heatmap "
            "--category site --series month --value ppm\n"
            "  validate_dataset.py data.xlsx --representation line "
            "--order week --y yield --sheet Results"
        ),
        formatter_class=argparse.RawDescriptionHelpFormatter,
    )
    parser.add_argument("path", type=Path, help="Path to a CSV or XLSX dataset")
    parser.add_argument(
        "--representation",
        required=True,
        choices=sorted(REPRESENTATIONS),
        help="Representation the activity asks students to produce",
    )
    for role in ROLE_FLAGS:
        parser.add_argument(
            f"--{role}",
            help=f"Column playing the {role} role in the requested representation",
        )
    parser.add_argument(
        "--column",
        action="append",
        default=[],
        help="Additional column the instructions name; repeat for each",
    )
    parser.add_argument(
        "--min-rows",
        type=int,
        help="Override the minimum usable observation count (must be positive)",
    )
    parser.add_argument("--sheet", help="Worksheet name for a multi-sheet workbook")
    parser.add_argument(
        "--json",
        action="store_true",
        help="Emit findings as JSON for programmatic callers",
    )
    return parser.parse_args()


def is_missing(value: str) -> bool:
    return normalize(value) in MISSING_TOKENS


def is_number(value: str) -> bool:
    try:
        number = float(value.replace(",", "").strip())
    except (ValueError, AttributeError):
        return False
    return math.isfinite(number)


def load_rows(
    path: Path, sheet: str | None = None
) -> tuple[list[str], list[list[str]]]:
    if not path.is_file():
        raise ValueError(f"File does not exist: {path}")
    suffix = path.suffix.lower()
    if suffix in {".xlsx", ".xlsm"}:
        try:
            # openpyxl is an optional dependency; CSV is the supported baseline.
            from openpyxl import load_workbook  # type: ignore[import-untyped]
        except ImportError as exc:  # pragma: no cover - environment dependent
            raise ValueError(
                f"Reading {suffix} requires openpyxl; convert the sheet to CSV "
                "or install openpyxl"
            ) from exc
        workbook = load_workbook(path, read_only=True, data_only=True)
        names = list(workbook.sheetnames)
        if sheet is not None:
            if sheet not in names:
                raise ValueError(
                    f"Worksheet not found: {sheet} (present: {', '.join(names)})"
                )
            worksheet = workbook[sheet]
        elif len(names) > 1:
            # Silently reading the active sheet of a multi-sheet workbook would
            # validate data the activity may not use at all.
            raise ValueError(
                f"Workbook has {len(names)} sheets ({', '.join(names)}); "
                "name one with --sheet"
            )
        else:
            worksheet = workbook[names[0]]
        records = [
            ["" if cell is None else str(cell) for cell in row]
            for row in worksheet.iter_rows(values_only=True)
        ]
        if not records:
            raise ValueError("Spreadsheet has no header row")
        return records[0], [
            row for row in records[1:] if any(cell.strip() for cell in row)
        ]
    with path.open(newline="", encoding="utf-8-sig") as handle:
        reader = csv.reader(handle)
        try:
            headers = next(reader)
        except StopIteration as exc:
            raise ValueError("Dataset has no header row") from exc
        rows = [row for row in reader if any(cell.strip() for cell in row)]
    return headers, rows


def column_values(headers: list[str], rows: list[list[str]], name: str) -> list[str]:
    index = headers.index(name)
    return [
        row[index] for row in rows if index < len(row) and not is_missing(row[index])
    ]


def classify(headers: list[str], rows: list[list[str]]) -> dict[str, str]:
    kinds: dict[str, str] = {}
    for header in headers:
        values = column_values(headers, rows, header)
        if not values:
            kinds[header] = "empty"
        elif all(is_number(value) for value in values):
            kinds[header] = QUANT
        else:
            kinds[header] = CATEG
    return kinds


def looks_like_identifier(header: str, values: list[str]) -> bool:
    """A numeric key is not a measurement, however numeric it looks.

    Two signals, deliberately ordered. A header naming a key settles it. A
    header naming a measured quantity settles it the other way, because the
    value-shape signal below cannot distinguish an index from a small-integer
    measurement and would otherwise reject real doses, times, and trial
    settings. Only a header that says neither falls through to value shape.
    """
    if IDENTIFIER_HEADER.search(header):
        return True
    if MEASUREMENT_HEADER.search(header):
        return False
    if len(values) < 3 or not all(is_number(value) for value in values):
        return False
    numbers = [float(value.replace(",", "")) for value in values]
    if not all(number.is_integer() for number in numbers):
        return False
    if len(set(numbers)) != len(numbers) or min(numbers) < 0:
        return False
    # Whole and unique is not enough — counts and totals are both. Neither is a
    # near-consecutive run on its own: consecutive integer measurements (20, 21,
    # 22, 23 °C) are a real column and were being rejected. Require the run to
    # also start where an index starts. A key numbered from some other origin
    # (1001..1004) is caught by the header instead, which is where such columns
    # almost always announce themselves.
    spread = max(numbers) - min(numbers)
    return min(numbers) in (0.0, 1.0) and spread <= 1.5 * (len(numbers) - 1)


def is_ordered(header: str, values: list[str]) -> bool:
    """An ordered key is a date or a number, or text a time-named header orders.

    Orderable is not the same as sorted: a line chart's points get sequenced
    before plotting, so unsorted weeks are fine. Dates and numbers carry their
    own order. Labels do not, so for text the header is the only evidence
    available — and it is evidence about the column, which is why it is
    consulted last and only when the values themselves settle nothing.
    """
    if not values:
        return False
    if all(parse_iso_date(value.strip()) for value in values):
        return True
    if all(is_number(value) for value in values):
        return True
    return bool(ORDERED_HEADER.search(header))


def complete_rows(headers: list[str], rows: list[list[str]], columns: list[str]) -> int:
    indexes = [headers.index(name) for name in columns if name in headers]
    if not indexes:
        return 0
    return sum(
        1
        for row in rows
        if all(index < len(row) and not is_missing(row[index]) for index in indexes)
    )


def validate(
    path: Path,
    representation: str,
    roles: dict[str, str | None],
    named_columns: list[str] | None = None,
    min_rows_override: int | None = None,
    sheet: str | None = None,
    strict_roles: bool = False,
) -> tuple[list[str], list[str]]:
    """Check a dataset against a representation's minimum requirements.

    `strict_roles` requires every role the representation defines, for every
    representation. The CLI leaves it off so exploratory checks stay cheap, but
    recorded evidence sets it: without named roles the permissive path only
    confirms that *some* column of each needed kind exists, which certifies the
    file rather than the variables the activity actually names.
    """
    if min_rows_override is not None and min_rows_override <= 0:
        return ["--min-rows must be a positive number of observations"], []
    try:
        headers, rows = load_rows(path, sheet)
    except (OSError, UnicodeError, ValueError) as exc:
        return [str(exc)], []

    issues: list[str] = []
    if not headers:
        return ["Dataset has no columns"], []
    normalized = [normalize(header) for header in headers]
    for name in sorted(
        {header for header in headers if normalized.count(normalize(header)) > 1}
    ):
        issues.append(f"Duplicate column header: {name}")
    if not rows:
        issues.append("Dataset has no data rows")

    lookup = {normalize(name): name for name in headers}
    kinds = classify(headers, rows)
    for name, kind in kinds.items():
        if kind == "empty":
            issues.append(f"Column has no usable values: {name}")

    def resolve(name: str, label: str) -> str | None:
        actual = lookup.get(normalize(name))
        if actual is None:
            issues.append(
                f"{label} column does not exist in the dataset: {name} "
                f"(present: {', '.join(headers)})"
            )
        return actual

    for name in named_columns or []:
        resolve(name, "Named")

    spec = REPRESENTATIONS[representation]
    supplied = {role: value for role, value in roles.items() if value}

    if spec.require_roles or strict_roles:
        missing = [role for role in spec.roles if role not in supplied]
        if missing:
            issues.append(
                f"{representation} requires explicit column roles; name "
                + ", ".join(f"--{role}" for role in sorted(missing))
                + ". Column types alone cannot tell a measurement from an identifier."
            )

    assigned: dict[str, str] = {}
    for role, expected in {**spec.roles, **spec.optional_roles}.items():
        declared = supplied.get(role)
        if not declared:
            continue
        actual = resolve(declared, f"--{role}")
        if actual is None:
            continue
        assigned[role] = actual
        values = column_values(headers, rows, actual)
        column_kind = kinds.get(actual)
        if expected == QUANT:
            if column_kind != QUANT:
                issues.append(
                    f"--{role} {actual} must be quantitative; "
                    f"its values are {column_kind}"
                )
            elif looks_like_identifier(actual, values):
                issues.append(
                    f"--{role} {actual} looks like an identifier or code rather than "
                    "a measurement; an identifier cannot carry the meaning this "
                    "representation assigns it"
                )
        elif expected == CATEG:
            if column_kind == QUANT and looks_like_identifier(actual, values):
                issues.append(
                    f"--{role} {actual} is a numeric key; confirm it names categories "
                    "rather than measurements"
                )
            levels = len({normalize(value) for value in values})
            if spec.min_levels and levels < spec.min_levels:
                issues.append(
                    f"--{role} {actual} has {levels} distinct level(s); "
                    f"{representation} needs at least {spec.min_levels} to mean anything"
                )
        elif expected == ORDER and not is_ordered(actual, values):
            issues.append(
                f"--{role} {actual} is not an ordered or time variable; "
                f"{representation} needs one to place points in sequence"
            )
        if role in spec.nonnegative:
            negatives = [
                value
                for value in values
                if is_number(value) and float(value.replace(",", "")) < 0
            ]
            if negatives:
                issues.append(
                    f"--{role} {actual} contains negative value(s) "
                    f"({', '.join(negatives[:3])}); {representation} shows parts of a "
                    "whole and cannot represent them"
                )
        if role in spec.unique:
            seen: set[str] = set()
            repeated_values: set[str] = set()
            for value in values:
                key = normalize(value)
                if key in seen:
                    repeated_values.add(value)
                seen.add(key)
            repeated = sorted(repeated_values)
            if repeated:
                issues.append(
                    f"--{role} {actual} repeats {', '.join(repeated[:3])}; "
                    f"{representation} needs mutually exclusive categories, so "
                    "repeated rows must be aggregated first"
                )

    # A column named for two roles produces a chart with one variable plotted
    # against itself, which the per-role type checks all pass.
    shared: dict[str, list[str]] = {}
    for role, actual in assigned.items():
        shared.setdefault(actual, []).append(role)
    for actual, roles_sharing in sorted(shared.items()):
        if len(roles_sharing) > 1:
            issues.append(
                f"{actual} is named for more than one role ("
                + ", ".join(f"--{role}" for role in sorted(roles_sharing))
                + f"); {representation} needs a distinct column for each"
            )

    # Without declared roles, fall back to counting kinds so the permissive
    # representations still get a check.
    if not spec.require_roles and not assigned:
        pool = [name for name in headers if kinds.get(name) != "empty"]
        need_quant = sum(1 for kind in spec.roles.values() if kind == QUANT)
        need_cat = sum(1 for kind in spec.roles.values() if kind == CATEG)
        quantitative = [name for name in pool if kinds.get(name) == QUANT]
        categorical = [name for name in pool if kinds.get(name) == CATEG]
        if len(quantitative) < need_quant:
            issues.append(
                f"{representation} needs {need_quant} quantitative column(s); found "
                f"{len(quantitative)} ({', '.join(quantitative) or 'none'})"
            )
        if len(categorical) < need_cat:
            issues.append(
                f"{representation} needs {need_cat} categorical column(s); found "
                f"{len(categorical)} ({', '.join(categorical) or 'none'})"
            )

    minimum = min_rows_override or spec.min_rows
    scope = sorted(set(assigned.values()))
    usable = complete_rows(headers, rows, scope) if scope else len(rows)
    if usable < minimum:
        label = (
            f"paired across {' and '.join(scope)}" if spec.paired and len(scope) >= 2
            else "usable"
        )
        issues.append(
            f"{representation} needs at least {minimum} {label} observation(s); "
            f"found {usable}"
        )

    return [], issues


def main() -> int:
    args = parse_args()
    errors, issues = validate(
        args.path,
        args.representation,
        {role: getattr(args, role) for role in ROLE_FLAGS},
        args.column,
        args.min_rows,
        args.sheet,
    )
    return emit_report(
        args.path,
        errors,
        issues,
        issue_label="GAP",
        ok_message=(
            f"dataset supports {args.representation}; whether the result answers "
            "the student's question still requires review"
        ),
        as_json=args.json,
    )


if __name__ == "__main__":
    sys.exit(main())
