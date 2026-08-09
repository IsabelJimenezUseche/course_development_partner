from __future__ import annotations

import hashlib
import importlib.util
import json
import re
import subprocess
import sys
import tempfile
import unittest
from pathlib import Path
from typing import NamedTuple

ROOT = Path(__file__).resolve().parents[1]
SCRIPTS = ROOT / "course-development-partner" / "scripts"
FIXTURES = ROOT / "tests" / "fixtures"


def run_script(
    name: str, content: str, suffix: str = ".md", *args: str
) -> subprocess.CompletedProcess[str]:
    with tempfile.TemporaryDirectory() as temp_dir:
        path = Path(temp_dir) / f"input{suffix}"
        path.write_text(content, encoding="utf-8")
        return subprocess.run(
            [sys.executable, str(SCRIPTS / name), str(path), *args],
            check=False,
            capture_output=True,
            text=True,
        )


def run_path(name: str, path: Path, *args: str) -> subprocess.CompletedProcess[str]:
    return subprocess.run(
        [sys.executable, str(SCRIPTS / name), str(path), *args],
        check=False,
        capture_output=True,
        text=True,
    )


def run_help(name: str) -> subprocess.CompletedProcess[str]:
    return subprocess.run(
        [sys.executable, str(SCRIPTS / name), "--help"],
        check=False,
        capture_output=True,
        text=True,
    )


def run_project(files: dict[str, str], *args: str) -> subprocess.CompletedProcess[str]:
    with tempfile.TemporaryDirectory() as temp_dir:
        project = Path(temp_dir)
        for name, content in files.items():
            if "- Schema version:" in content and "- Last updated:" not in content:
                content = re.sub(
                    r"(^\s*-\s*Schema version:\s*\S+\s*$)",
                    r"\1\n- Last updated: 2026-08-01",
                    content,
                    count=1,
                    flags=re.MULTILINE,
                )
            path = project / name
            path.parent.mkdir(parents=True, exist_ok=True)
            path.write_text(content, encoding="utf-8")
        return subprocess.run(
            [sys.executable, str(SCRIPTS / "validate_project.py"), str(project), *args],
            check=False,
            capture_output=True,
            text=True,
        )


class DesignStateValidatorTests(unittest.TestCase):
    def test_complete_design_state_passes(self) -> None:
        content = """# Course Design Brief
- Engagement tier: Project
## Course context
- Course or module: BIO 101
## Intended learning
- Learning outcomes: Explain a mechanism
## Access, participation, and belonging
- Provide equivalent text and visual representations
## Constraints
- Fifty minutes
## Implementation load
- One class period; no new platform
## Collaboration
- Interaction level: Co-design
- Requested artifacts: Alignment map and activity draft
## Status
### Confirmed
- Outcome approved
### Assumed
- Prior knowledge present
### Open
- Accessibility review pending
### Current phase
- Establish alignment
### Next decision
- Select evidence of learning
"""
        result = run_script("validate_design_state.py", content)
        self.assertEqual(result.returncode, 0, result.stdout + result.stderr)

    def test_blank_field_does_not_hide_the_field_below_it(self) -> None:
        """A blank bullet used to swallow the next line, hiding a filled field.

        `\\s` matches newlines and MULTILINE `$` closes at any line end, so the scan
        ran past a blank field and never saw the one beneath it. Real briefs are
        part-filled, so this misfired exactly when it mattered most.
        """
        content = (FIXTURES / "design_state" / "valid.md").read_text(encoding="utf-8")
        # Blank the field directly above a profile-required one.
        content = content.replace(
            "- Engagement tier:", "- Last updated:\n- Engagement tier:"
        )
        result = run_script("validate_design_state.py", content)

        # `Engagement tier` is filled and sits under a blank field; it must be seen.
        self.assertNotIn(
            "unanswered for establish: engagement tier", result.stdout.lower()
        )

    def test_profile_decides_the_verdict_not_every_blank_field(self) -> None:
        """An instructor at `establish` is not blocked by a later phase's field.

        `Minimum viable fallback` is required at produce and handoff and not at
        establish, so the same brief must clear establish and be held at the others.
        Before this, any blank anywhere failed all three and `--profile` could not
        change a verdict.
        """
        brief = (FIXTURES / "design_state" / "valid.md").read_text(encoding="utf-8")
        blanked = re.sub(
            r"^- Minimum viable fallback:.*$",
            "- Minimum viable fallback:",
            brief,
            flags=re.MULTILINE,
        )
        self.assertNotEqual(brief, blanked, "fixture lacks the field this test needs")

        establish = run_script("validate_design_state.py", blanked, ".md", "--profile", "establish")
        produce = run_script("validate_design_state.py", blanked, ".md", "--profile", "produce")

        self.assertEqual(establish.returncode, 0, establish.stdout + establish.stderr)
        self.assertIn("NOTE:", establish.stdout)
        self.assertEqual(produce.returncode, 2, produce.stdout + produce.stderr)
        self.assertIn("minimum viable fallback", produce.stdout.lower())

    def test_advisory_notes_never_change_the_exit_code(self) -> None:
        sys.path.insert(0, str(SCRIPTS))
        try:
            from _tabular import emit_report
        finally:
            sys.path.pop(0)

        code = emit_report(
            "x.md", [], [], issue_label="INCOMPLETE", ok_message="fine",
            notes=["Unanswered field on line 9: something later"],
        )
        self.assertEqual(code, 0)

    def test_unanswered_fields_are_reported_with_a_location(self) -> None:
        """A bare 'a field is unanswered' is unactionable in a 67-field template."""
        template = (
            ROOT / "course-development-partner" / "assets" / "course-design-brief.md"
        ).read_text()
        result = run_script("validate_design_state.py", template)

        self.assertEqual(result.returncode, 2, result.stdout + result.stderr)
        self.assertRegex(result.stdout, r"Unanswered field on line \d+: \S")

    def test_blank_scalar_does_not_capture_the_next_line_as_its_value(self) -> None:
        """A blank `Evidence level claimed:` used to read the following bullet.

        Asserted against the extractor directly. Driving it through the CLI proved
        nothing: the validator reports other problems first, so the wrong value
        never reached stdout and the test passed against the bug.
        """
        sys.path.insert(0, str(SCRIPTS))
        try:
            import validate_assessment_blueprint as blueprint_validator
        finally:
            sys.path.pop(0)

        text = "- Evidence level claimed:\n- Known limitations: none\n"
        value = blueprint_validator.evidence_level(text, [], {})

        self.assertEqual(value, "", f"blank field captured {value!r} from the next line")

    def test_incomplete_design_state_returns_two(self) -> None:
        template = (
            ROOT / "course-development-partner" / "assets" / "course-design-brief.md"
        ).read_text()
        result = run_script("validate_design_state.py", template)
        self.assertEqual(result.returncode, 2, result.stdout + result.stderr)

    def test_required_heading_at_wrong_level_is_structural_error(self) -> None:
        content = (FIXTURES / "design_state" / "valid.md").read_text(encoding="utf-8")
        content = content.replace("## Constraints", "### Constraints")
        result = run_script("validate_design_state.py", content)
        self.assertEqual(result.returncode, 1, result.stdout + result.stderr)
        self.assertIn("Missing required level-2 heading: constraints", result.stdout)

    def test_duplicate_scalar_field_is_structural_error(self) -> None:
        content = (FIXTURES / "design_state" / "valid.md").read_text(encoding="utf-8")
        content = content.replace(
            "- Engagement tier: Project",
            "- Engagement tier: Project\n- Engagement tier: Course",
        )
        result = run_script("validate_design_state.py", content)
        self.assertEqual(result.returncode, 1, result.stdout + result.stderr)
        self.assertIn("Duplicate scalar field: engagement tier", result.stdout)

    def test_legitimate_bracketed_content_is_not_a_placeholder(self) -> None:
        content = (FIXTURES / "design_state" / "valid.md").read_text(encoding="utf-8")
        content = content.replace(
            "Explain a mechanism",
            "Interpret a confidence interval such as [0, 1]",
        )
        result = run_script("validate_design_state.py", content)
        self.assertEqual(result.returncode, 0, result.stdout + result.stderr)


class MachineReadableReportTests(unittest.TestCase):
    """A caller driving these in a loop should read findings, not parse prose."""

    VALIDATORS = (
        ("validate_alignment_map.py", "alignment"),
        ("validate_assessment_blueprint.py", "assessment_blueprint"),
        ("validate_artifact_manifest.py", "artifact_manifest"),
        ("validate_course_curriculum_map.py", "course_curriculum_map"),
        ("validate_design_state.py", "design_state"),
    )

    def test_every_validator_emits_json_matching_its_exit_code(self) -> None:
        for script, fixture_dir in self.VALIDATORS:
            for name, expected in (
                ("valid", "pass"),
                ("invalid", "incomplete"),
                ("malformed", "fail"),
            ):
                path = FIXTURES / fixture_dir / f"{name}.md"
                if not path.is_file():
                    continue
                with self.subTest(script=script, fixture=name):
                    result = run_path(script, path, "--json")
                    payload = json.loads(result.stdout)
                    # The exit code and the reported status must never disagree.
                    self.assertEqual(payload["exit_code"], result.returncode)
                    self.assertEqual(payload["status"], expected)
                    if expected != "pass":
                        self.assertTrue(
                            payload["findings"], "a failing run must report findings"
                        )
                        for finding in payload["findings"]:
                            self.assertIn("level", finding)
                            self.assertIn("message", finding)

    def test_json_flag_does_not_change_the_verdict(self) -> None:
        for script, fixture_dir in self.VALIDATORS:
            path = FIXTURES / fixture_dir / "invalid.md"
            if not path.is_file():
                continue
            with self.subTest(script=script):
                self.assertEqual(
                    run_path(script, path).returncode,
                    run_path(script, path, "--json").returncode,
                )


class LookalikeCharacterTests(unittest.TestCase):
    """Typographic look-alikes must not decide whether a valid file is accepted.

    Generated Markdown routinely carries a non-breaking hyphen or an en dash where
    ASCII was meant. A reader cannot see the difference, so rejecting the file
    teaches nothing and the author has no way to find the cause.
    """

    def test_non_breaking_hyphen_in_a_column_heading_is_accepted(self) -> None:
        """The exact failure seen in production: `Construct-irrelevant barriers`
        written with U+2011 made the whole blueprint unparseable."""
        blueprint = (FIXTURES / "assessment_blueprint" / "valid.md").read_text(
            encoding="utf-8"
        )
        self.assertIn("Construct-irrelevant barriers", blueprint)
        # Swap only the ASCII hyphen for U+2011 NON-BREAKING HYPHEN.
        hyphenated = blueprint.replace(
            "Construct-irrelevant barriers", "Construct‑irrelevant barriers"
        )
        self.assertNotEqual(blueprint, hyphenated)

        control = run_script("validate_assessment_blueprint.py", blueprint)
        result = run_script("validate_assessment_blueprint.py", hyphenated)

        self.assertEqual(control.returncode, 0, control.stdout + control.stderr)
        self.assertEqual(
            result.returncode,
            control.returncode,
            f"one look-alike character changed the verdict:\n{result.stdout}",
        )

    def test_en_dash_in_a_controlled_token_is_still_rejected(self) -> None:
        """Folding must not turn an unknown token into a valid one."""
        sys.path.insert(0, str(SCRIPTS))
        try:
            from _tabular import fold_lookalikes, normalize, parse_cognitive_demand
        finally:
            sys.path.pop(0)

        self.assertEqual(fold_lookalikes("Construct‑irrelevant"), "Construct-irrelevant")
        self.assertEqual(normalize("Construct‑irrelevant"), "construct-irrelevant")
        # A look-alike inside a real word is folded, but nonsense stays nonsense.
        self.assertIsNone(parse_cognitive_demand("analyse‑ish"))
        self.assertEqual(parse_cognitive_demand("Analyze"), 4)


class AlignmentValidatorTests(unittest.TestCase):
    def test_complete_alignment_passes(self) -> None:
        content = """# Alignment Map
| Outcome ID | Observable learning outcome | Cognitive demand | Evidence of learning | Learning mechanism | Learning activity/support | Feedback or assessment | Status |
|---|---|---|---|---|---|---|---|
| LO-1 | Explain a mechanism | Understand | Annotated diagram | contrasting cases | Comparison activity | Exit ticket | approved |
"""
        result = run_script("validate_alignment_map.py", content)
        self.assertEqual(result.returncode, 0, result.stdout + result.stderr)

    def test_alignment_gap_returns_two(self) -> None:
        content = """| Outcome ID | Observable learning outcome | Cognitive demand | Evidence of learning | Learning mechanism | Learning activity/support | Feedback or assessment | Status |
|---|---|---|---|---|---|---|---|
| LO-1 | Explain a mechanism | understand |  | contrasting cases | Comparison activity | Exit ticket | draft |
"""
        result = run_script("validate_alignment_map.py", content)
        self.assertEqual(result.returncode, 2, result.stdout + result.stderr)
        self.assertIn("missing evidence of learning", result.stdout)

    def test_empty_and_unknown_statuses_are_both_rejected(self) -> None:
        result = run_path(
            "validate_alignment_map.py", FIXTURES / "alignment" / "invalid.md"
        )
        self.assertEqual(result.returncode, 2, result.stdout + result.stderr)
        self.assertIn("missing status", result.stdout)
        self.assertIn("unknown status done", result.stdout)
        self.assertIn("duplicate outcome ID LO-1", result.stdout)
        self.assertIn("activity or assessment has no outcome ID", result.stdout)

    def test_missing_alignment_fields_and_empty_row_are_reported(self) -> None:
        content = """| Outcome ID | Observable learning outcome | Cognitive demand | Evidence of learning | Learning mechanism | Learning activity/support | Feedback or assessment | Status |
|---|---|---|---|---|---|---|---|
| LO-1 |  |  |  |  |  |  |  |
|  |  |  |  |  |  |  |  |
"""
        result = run_script("validate_alignment_map.py", content)
        self.assertEqual(result.returncode, 2, result.stdout + result.stderr)
        for expected in (
            "missing observable learning outcome",
            "missing cognitive demand",
            "missing evidence of learning",
            "missing learning mechanism",
            "missing learning activity/support",
            "missing feedback or assessment",
            "missing status",
            "empty row",
        ):
            self.assertIn(expected, result.stdout)

    def test_parser_selects_schema_table_and_preserves_escaped_pipe(self) -> None:
        content = """| Note | Value |
|---|---|
| Owner | Faculty |

| Outcome ID | Observable learning outcome | Cognitive demand | Evidence of learning | Learning mechanism | Learning activity/support | Feedback or assessment | Status |
|---|---|---|---|---|---|---|---|
| LO-1 | Compare A \\| B | analyze | Explanation | contrasting cases | Contrast cases | Feedback | approved |
"""
        result = run_script("validate_alignment_map.py", content)
        self.assertEqual(result.returncode, 0, result.stdout + result.stderr)

    def test_uncontrolled_cognitive_demand_is_reported(self) -> None:
        content = """| Outcome ID | Observable learning outcome | Cognitive demand | Evidence of learning | Learning mechanism | Learning activity/support | Feedback or assessment | Status |
|---|---|---|---|---|---|---|---|
| LO-1 | Explain a mechanism | synthesize | Diagram | contrasting cases | Activity | Exit ticket | approved |
"""
        result = run_script("validate_alignment_map.py", content)
        self.assertEqual(result.returncode, 2, result.stdout + result.stderr)
        self.assertIn("unknown cognitive demand synthesize", result.stdout)

    def test_every_controlled_cognitive_demand_is_accepted(self) -> None:
        rows = "\n".join(
            f"| LO-{index} | Outcome {index} | {demand} | Evidence | retrieval | Activity | Feedback | approved |"
            for index, demand in enumerate(
                ("remember", "understand", "apply", "analyze", "evaluate", "create"),
                start=1,
            )
        )
        content = (
            "| Outcome ID | Observable learning outcome | Cognitive demand | Evidence of learning "
            "| Learning mechanism | Learning activity/support | Feedback or assessment | Status |\n"
            "|---|---|---|---|---|---|---|---|\n" + rows + "\n"
        )
        result = run_script("validate_alignment_map.py", content)
        self.assertEqual(result.returncode, 0, result.stdout + result.stderr)

    def test_malformed_middle_row_is_structural_error(self) -> None:
        content = """| Outcome ID | Observable learning outcome | Cognitive demand | Evidence of learning | Learning mechanism | Learning activity/support | Feedback or assessment | Status |
|---|---|---|---|---|---|---|---|
| LO-1 | Explain | understand | Evidence | retrieval | Activity | Feedback | approved |
| LO-2 | This row is short | Evidence |
| LO-3 | Apply | apply | Evidence | practice | Activity | Feedback | approved |
"""
        result = run_script("validate_alignment_map.py", content)
        self.assertEqual(result.returncode, 1, result.stdout + result.stderr)
        self.assertIn("Malformed Markdown table row", result.stdout)

    def test_duplicate_normalized_header_is_structural_error(self) -> None:
        content = """| Outcome ID | outcome   id | Observable learning outcome | Cognitive demand | Evidence of learning | Learning mechanism | Learning activity/support | Feedback or assessment | Status |
|---|---|---|---|---|---|---|---|---|
| LO-1 | LO-1 | Explain | understand | Evidence | retrieval | Activity | Feedback | approved |
"""
        result = run_script("validate_alignment_map.py", content)
        self.assertEqual(result.returncode, 1, result.stdout + result.stderr)
        self.assertIn("Duplicate normalized header", result.stdout)

    def test_retired_only_alignment_reports_no_active_outcomes(self) -> None:
        content = """| Outcome ID | Observable learning outcome | Cognitive demand | Evidence of learning | Learning mechanism | Learning activity/support | Feedback or assessment | Status |
|---|---|---|---|---|---|---|---|
| LO-1 | Explain | understand | Explanation | retrieval | Archived activity | Archived feedback | retired |
"""
        result = run_script("validate_alignment_map.py", content)
        self.assertEqual(result.returncode, 2, result.stdout + result.stderr)
        self.assertIn("Alignment map contains no active outcomes", result.stdout)

    def test_status_only_retired_row_requires_an_outcome_id(self) -> None:
        content = """| Outcome ID | Observable learning outcome | Cognitive demand | Evidence of learning | Learning mechanism | Learning activity/support | Feedback or assessment | Status |
|---|---|---|---|---|---|---|---|
|  |  |  |  |  |  |  | retired |
| LO-1 | Explain | understand | Explanation | retrieval | Activity | Feedback | approved |
"""
        result = run_script("validate_alignment_map.py", content)
        self.assertEqual(result.returncode, 2, result.stdout + result.stderr)
        self.assertIn("activity or assessment has no outcome ID", result.stdout)

    def test_separator_only_outcome_id_is_reported_without_a_traceback(self) -> None:
        content = """| Outcome ID | Observable learning outcome | Cognitive demand | Evidence of learning | Learning mechanism | Learning activity/support | Feedback or assessment | Status |
|---|---|---|---|---|---|---|---|
| ; | Explain | understand | Explanation | retrieval | Activity | Feedback | approved |
"""
        result = run_script("validate_alignment_map.py", content)
        self.assertEqual(result.returncode, 2, result.stdout + result.stderr)
        self.assertIn("identifier list contains no identifiers", result.stdout)
        self.assertNotIn("Traceback", result.stderr)


class ArtifactManifestValidatorTests(unittest.TestCase):
    def test_teaching_ready_manifest_passes(self) -> None:
        content = """| Artifact ID | Artifact type | Artifact family | Variant | Required variants | File or reference | Audience | Outcome(s) | Status | Validation completed | Blockers/open issues | Last reviewed | Production plan | Accessibility review | Safety review | Data-task-fit evidence |
|---|---|---|---|---|---|---|---|---|---|---|---|---|---|---|---|
| WS-1 | markdown |  |  |  | https://example.edu/ws-1 | student | LO-1 | teaching-ready | technical; alignment; accessibility; reopen | none | 2026-07-31 | not required | https://example.edu/ws-1-accessibility | not required | not applicable — no dataset |
"""
        result = run_script("validate_artifact_manifest.py", content)
        self.assertEqual(result.returncode, 0, result.stdout + result.stderr)

    def test_teaching_ready_with_blocker_returns_two(self) -> None:
        content = """| Artifact ID | Artifact type | Artifact family | Variant | Required variants | File or reference | Audience | Outcome(s) | Status | Validation completed | Blockers/open issues | Last reviewed | Production plan | Accessibility review | Safety review | Data-task-fit evidence |
|---|---|---|---|---|---|---|---|---|---|---|---|---|---|---|---|
| WS-1 | markdown |  |  |  | worksheet.md | student | LO-1 | teaching-ready | technical; alignment; accessibility; reopen | solution not verified | 2026-07-31 | not required | accessibility-review.md | not required | not applicable — no dataset |
"""
        result = run_script("validate_artifact_manifest.py", content)
        self.assertEqual(result.returncode, 2, result.stdout + result.stderr)
        self.assertIn("unresolved blockers/open issues", result.stdout)

    def test_declared_pair_requires_labeled_student_and_instructor_variants(
        self,
    ) -> None:
        result = run_path(
            "validate_artifact_manifest.py",
            FIXTURES / "artifact_manifest" / "invalid.md",
        )
        self.assertEqual(result.returncode, 2, result.stdout + result.stderr)
        self.assertIn("missing a variant label", result.stdout)
        self.assertIn("required instructor variant is not represented", result.stdout)

    def test_complete_declared_pair_passes(self) -> None:
        result = run_path(
            "validate_artifact_manifest.py",
            FIXTURES / "artifact_manifest" / "boundary.md",
        )
        self.assertEqual(result.returncode, 0, result.stdout + result.stderr)

    def test_check_paths_detects_missing_local_reference(self) -> None:
        content = """| Artifact ID | Artifact type | Artifact family | Variant | Required variants | File or reference | Audience | Outcome(s) | Status | Validation completed | Blockers/open issues | Last reviewed | Production plan | Accessibility review | Safety review | Data-task-fit evidence |
|---|---|---|---|---|---|---|---|---|---|---|---|---|---|---|---|
| WS-1 | markdown |  |  |  | missing.md | student | LO-1 | draft | manual | none | 2026-07-31 | https://example.edu/not-required | https://example.edu/pending | not required | not applicable — no dataset |
"""
        result = run_script(
            "validate_artifact_manifest.py", content, ".md", "--check-paths"
        )
        self.assertEqual(result.returncode, 2, result.stdout + result.stderr)
        self.assertIn("local file or reference does not exist", result.stdout)

    def test_check_paths_reports_a_symlink_loop_without_a_traceback(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            directory = Path(temp_dir)
            (directory / "loop.md").symlink_to("loop.md")
            manifest = directory / "artifact-manifest.md"
            manifest.write_text(
                """| Artifact ID | Artifact type | Artifact family | Variant | Required variants | File or reference | Audience | Outcome(s) | Status | Validation completed | Blockers/open issues | Last reviewed | Production plan | Accessibility review | Safety review | Data-task-fit evidence |
|---|---|---|---|---|---|---|---|---|---|---|---|---|---|---|---|
| WS-1 | markdown |  |  |  | loop.md | student | LO-1 | draft | manual | none | 2026-08-01 | not required | pending | not required | not applicable — no dataset |
""",
                encoding="utf-8",
            )
            result = run_path(
                "validate_artifact_manifest.py",
                manifest,
                "--check-paths",
            )
        self.assertEqual(result.returncode, 1, result.stdout + result.stderr)
        self.assertIn("cannot resolve local file or reference", result.stdout)
        self.assertNotIn("Traceback", result.stderr)

    def test_missing_fields_validation_evidence_and_family_context_are_reported(
        self,
    ) -> None:
        content = """| Artifact ID | Artifact family | Variant | Required variants | Artifact type | File or reference | Audience | Outcome(s) | Status | Validation completed | Blockers/open issues | Last reviewed | Production plan | Accessibility review | Safety review | Data-task-fit evidence |
|---|---|---|---|---|---|---|---|---|---|---|---|---|---|---|---|
|  |  | student | student; instructor | markdown |  |  |  |  |  | none |  | not required | pending | not required | not applicable — no dataset |
| A-2 |  |  |  | markdown | https://example.edu/a-2 | instructor | LO-1 | validated |  | none | 2026-07-31 | not required | pending | not required | not applicable — no dataset |
"""
        result = run_script("validate_artifact_manifest.py", content)
        self.assertEqual(result.returncode, 2, result.stdout + result.stderr)
        for expected in (
            "missing artifact ID",
            "missing file or reference",
            "missing audience",
            "missing outcome(s)",
            "missing status",
            "validated without validation evidence",
            "variant or required variants declared without an artifact family",
        ):
            self.assertIn(expected, result.stdout)

    def test_ready_manifest_rejects_none_evidence_and_invalid_date(self) -> None:
        content = """| Artifact ID | Artifact type | Artifact family | Variant | Required variants | File or reference | Audience | Outcome(s) | Status | Validation completed | Blockers/open issues | Last reviewed | Production plan | Accessibility review | Safety review | Data-task-fit evidence |
|---|---|---|---|---|---|---|---|---|---|---|---|---|---|---|---|
| WS-1 | markdown |  |  |  | https://example.edu/ws | student | LO-1 | validated | none | none | 2026-13-40 | not required | pending | not required | not applicable — no dataset |
"""
        result = run_script("validate_artifact_manifest.py", content)
        self.assertEqual(result.returncode, 2, result.stdout + result.stderr)
        self.assertIn("unknown validation token none", result.stdout)
        self.assertIn("last reviewed must use YYYY-MM-DD", result.stdout)

    def test_calendar_invalid_iso_date_is_rejected(self) -> None:
        content = """| Artifact ID | Artifact type | Artifact family | Variant | Required variants | File or reference | Audience | Outcome(s) | Status | Validation completed | Blockers/open issues | Last reviewed | Production plan | Accessibility review | Safety review | Data-task-fit evidence |
|---|---|---|---|---|---|---|---|---|---|---|---|---|---|---|---|
| WS-1 | markdown |  |  |  | https://example.edu/ws | student | LO-1 | validated | technical; alignment | none | 2026-02-31 | not required | pending | not required | not applicable — no dataset |
"""
        result = run_script("validate_artifact_manifest.py", content)
        self.assertEqual(result.returncode, 2, result.stdout + result.stderr)
        self.assertIn("last reviewed must use YYYY-MM-DD", result.stdout)

    def test_draft_manifest_rejects_a_non_iso_review_date(self) -> None:
        content = """| Artifact ID | Artifact type | Artifact family | Variant | Required variants | File or reference | Audience | Outcome(s) | Status | Validation completed | Blockers/open issues | Last reviewed | Production plan | Accessibility review | Safety review | Data-task-fit evidence |
|---|---|---|---|---|---|---|---|---|---|---|---|---|---|---|---|
| WS-1 | markdown |  |  |  | https://example.edu/ws | student | LO-1 | draft | manual | none | not-a-date | not required | pending | not required | not applicable — no dataset |
"""
        result = run_script("validate_artifact_manifest.py", content)
        self.assertEqual(result.returncode, 2, result.stdout + result.stderr)
        self.assertIn("last reviewed must use YYYY-MM-DD", result.stdout)

    def test_pending_non_file_evidence_is_not_treated_as_a_path_for_drafts(
        self,
    ) -> None:
        content = """| Artifact ID | Artifact type | Artifact family | Variant | Required variants | File or reference | Audience | Outcome(s) | Status | Validation completed | Blockers/open issues | Last reviewed | Production plan | Accessibility review | Safety review | Data-task-fit evidence |
|---|---|---|---|---|---|---|---|---|---|---|---|---|---|---|---|
| WS-1 | markdown |  |  |  | https://example.edu/ws | student | LO-1 | draft | manual | none | 2026-08-01 | not applicable — plain Markdown | pending | not required | not applicable — no dataset |
"""
        result = run_script(
            "validate_artifact_manifest.py", content, ".md", "--check-paths"
        )
        self.assertEqual(result.returncode, 0, result.stdout + result.stderr)

    def test_every_declared_variant_is_required(self) -> None:
        content = """| Artifact ID | Artifact family | Variant | Required variants | Artifact type | File or reference | Audience | Outcome(s) | Status | Validation completed | Blockers/open issues | Last reviewed | Production plan | Accessibility review | Safety review | Data-task-fit evidence |
|---|---|---|---|---|---|---|---|---|---|---|---|---|---|---|---|
| WS-1 | family-1 | student | student; instructor; solution | markdown | https://example.edu/student | student | LO-1 | draft | manual | none | 2026-07-31 | not required | pending | not required | not applicable — no dataset |
| WS-2 | family-1 | instructor | student; instructor; solution | markdown | https://example.edu/instructor | instructor | LO-1 | draft | manual | none | 2026-07-31 | not required | pending | not required | not applicable — no dataset |
"""
        result = run_script("validate_artifact_manifest.py", content)
        self.assertEqual(result.returncode, 2, result.stdout + result.stderr)
        self.assertIn("required solution variant is not represented", result.stdout)

    def test_manifest_requires_artifact_family_schema_columns(self) -> None:
        content = """| Artifact ID | Artifact type | File or reference | Audience | Outcome(s) | Status | Validation completed | Blockers/open issues | Last reviewed | Production plan | Accessibility review | Safety review | Data-task-fit evidence |
|---|---|---|---|---|---|---|---|---|---|---|---|---|
| WS-1 | markdown | https://example.edu/ws | student | LO-1 | draft | manual | none | 2026-08-01 | not required | pending | not required | not applicable — no dataset |
"""
        result = run_script("validate_artifact_manifest.py", content)
        self.assertEqual(result.returncode, 1, result.stdout + result.stderr)
        self.assertIn("No Markdown table contains the required columns", result.stdout)

    MANIFEST_HEADER = (
        "| Artifact ID | Artifact type | Artifact family | Variant | Required variants "
        "| File or reference | Audience | Outcome(s) | Status | Validation completed "
        "| Blockers/open issues | Last reviewed | Production plan | Accessibility review "
        "| Safety review | Data-task-fit evidence |\n"
        "|---|---|---|---|---|---|---|---|---|---|---|---|---|---|---|---|\n"
    )

    def _teaching_ready_row(
        self,
        safety: str,
        fit_evidence: str = "not applicable — no dataset",
        tokens: str = "technical; alignment; accessibility; reopen",
    ) -> str:
        return (
            self.MANIFEST_HEADER
            + "| WS-1 | markdown |  |  |  | https://example.edu/ws-1 | student | LO-1 "
            f"| teaching-ready | {tokens} | none "
            f"| 2026-07-31 | not required | https://example.edu/a11y | {safety} "
            f"| {fit_evidence} |\n"
        )

    def test_teaching_ready_without_safety_declaration_is_reported(self) -> None:
        result = run_script(
            "validate_artifact_manifest.py", self._teaching_ready_row("")
        )
        self.assertEqual(result.returncode, 2, result.stdout + result.stderr)
        self.assertIn("no safety review declaration", result.stdout)

    def test_teaching_ready_accepts_explicit_not_required_safety(self) -> None:
        result = run_script(
            "validate_artifact_manifest.py", self._teaching_ready_row("not required")
        )
        self.assertEqual(result.returncode, 0, result.stdout + result.stderr)

    def test_teaching_ready_accepts_a_linked_safety_review(self) -> None:
        result = run_script(
            "validate_artifact_manifest.py",
            self._teaching_ready_row("[safety](https://example.edu/safety-review)"),
        )
        self.assertEqual(result.returncode, 0, result.stdout + result.stderr)

    def test_teaching_ready_distinguishes_prose_from_a_local_review_path(
        self,
    ) -> None:
        prose_result = run_script(
            "validate_artifact_manifest.py",
            self._teaching_ready_row("Approved by EHS. See lab binder"),
        )
        self.assertEqual(
            prose_result.returncode,
            2,
            prose_result.stdout + prose_result.stderr,
        )
        self.assertIn(
            "must be a review reference",
            prose_result.stdout,
        )

        path_result = run_script(
            "validate_artifact_manifest.py",
            self._teaching_ready_row("safety reviews/lab approval.pdf"),
        )
        self.assertEqual(
            path_result.returncode,
            0,
            path_result.stdout + path_result.stderr,
        )

    def test_teaching_ready_rejects_pending_or_bare_safety_status(self) -> None:
        for safety in ("pending", "approved"):
            with self.subTest(safety=safety):
                result = run_script(
                    "validate_artifact_manifest.py",
                    self._teaching_ready_row(safety),
                )
                self.assertEqual(result.returncode, 2, result.stdout + result.stderr)
                self.assertIn("teaching-ready", result.stdout)

    def test_manifest_rejects_invalid_artifact_id(self) -> None:
        content = self._teaching_ready_row("not required").replace(
            "| WS-1 | markdown", "| not a valid id | markdown"
        )
        result = run_script("validate_artifact_manifest.py", content)
        self.assertEqual(result.returncode, 2, result.stdout + result.stderr)
        self.assertIn("invalid artifact identifier", result.stdout)

    def test_draft_row_may_leave_safety_review_blank(self) -> None:
        content = (
            self.MANIFEST_HEADER
            + "| WS-1 | markdown |  |  |  | https://example.edu/ws-1 | student | LO-1 "
            "| draft | manual | none | 2026-07-31 | not required | pending |  |  |\n"
        )
        result = run_script("validate_artifact_manifest.py", content)
        self.assertEqual(result.returncode, 0, result.stdout + result.stderr)

    READY_TOKENS = "technical; alignment; accessibility; reopen"
    FIT_TOKENS = f"{READY_TOKENS}; data-task-fit"

    def test_data_task_fit_token_without_a_record_is_rejected(self) -> None:
        """The token asserts executed work; alone it is unverifiable.

        This is the self-attestation hole: the earlier schema let a row carry
        the token with nothing behind it, so a row that did the work and a row
        that typed the word were indistinguishable.
        """
        content = self._teaching_ready_row(
            "not required", "not applicable — no dataset", self.FIT_TOKENS
        )
        result = run_script("validate_artifact_manifest.py", content)
        self.assertEqual(result.returncode, 2, result.stdout + result.stderr)
        self.assertIn("without referencing the record", result.stdout)

    def test_linked_record_without_the_token_is_rejected(self) -> None:
        content = self._teaching_ready_row("not required", "data-task-record.md")
        result = run_script("validate_artifact_manifest.py", content)
        self.assertEqual(result.returncode, 2, result.stdout + result.stderr)
        self.assertIn("missing the data-task-fit validation token", result.stdout)

    def test_linked_record_with_the_token_passes(self) -> None:
        content = self._teaching_ready_row(
            "not required", "data-task-record.md", self.FIT_TOKENS
        )
        result = run_script("validate_artifact_manifest.py", content)
        self.assertEqual(result.returncode, 0, result.stdout + result.stderr)

    def test_undeclared_data_task_fit_is_reported(self) -> None:
        content = self._teaching_ready_row("not required", "")
        result = run_script("validate_artifact_manifest.py", content)
        self.assertEqual(result.returncode, 2, result.stdout + result.stderr)
        self.assertIn("no data-task-fit declaration", result.stdout)

    def test_pending_data_task_fit_cannot_support_teaching_ready(self) -> None:
        content = self._teaching_ready_row("not required", "pending")
        result = run_script("validate_artifact_manifest.py", content)
        self.assertEqual(result.returncode, 2, result.stdout + result.stderr)
        self.assertIn("unresolved data-task-fit evidence state", result.stdout)

    def test_prose_is_not_a_data_task_fit_record(self) -> None:
        content = self._teaching_ready_row("not required", "I checked the columns")
        result = run_script("validate_artifact_manifest.py", content)
        self.assertEqual(result.returncode, 2, result.stdout + result.stderr)
        self.assertIn("must reference the record", result.stdout)

    def test_retired_variant_does_not_satisfy_active_family_requirement(self) -> None:
        content = """| Artifact ID | Artifact type | Artifact family | Variant | Required variants | File or reference | Audience | Outcome(s) | Status | Validation completed | Blockers/open issues | Last reviewed | Production plan | Accessibility review | Safety review | Data-task-fit evidence |
|---|---|---|---|---|---|---|---|---|---|---|---|---|---|---|---|
| WS-S | markdown | worksheet | student | student; instructor; solution | https://example.edu/student | student | LO-1 | draft | manual | none | 2026-08-01 | not required | pending | not required | not applicable — no dataset |
| WS-I | markdown | worksheet | instructor |  | https://example.edu/instructor | instructor | LO-1 | draft | manual | none | 2026-08-01 | not required | pending | not required | not applicable — no dataset |
| WS-K | markdown | worksheet | solution |  | old-solution.md | instructor | LO-1 | retired | manual | none | 2026-07-01 | not required | pending | not required | not applicable — no dataset |
"""
        result = run_script(
            "validate_artifact_manifest.py",
            content,
            ".md",
            "--check-paths",
        )
        self.assertEqual(result.returncode, 2, result.stdout + result.stderr)
        self.assertIn("required solution variant is not represented", result.stdout)
        self.assertNotIn("old-solution.md", result.stdout)

    def test_retired_only_manifest_reports_no_active_artifacts(self) -> None:
        content = """| Artifact ID | Artifact type | Artifact family | Variant | Required variants | File or reference | Audience | Outcome(s) | Status | Validation completed | Blockers/open issues | Last reviewed | Production plan | Accessibility review | Safety review | Data-task-fit evidence |
|---|---|---|---|---|---|---|---|---|---|---|---|---|---|---|---|
| OLD-1 | markdown |  |  |  | old.md | instructor | LO-1 | retired | manual | none | 2026-07-01 | not required | pending | not required | not applicable — no dataset |
"""
        result = run_script("validate_artifact_manifest.py", content)
        self.assertEqual(result.returncode, 2, result.stdout + result.stderr)
        self.assertIn("Artifact manifest contains no active artifacts", result.stdout)


class AssessmentBlueprintValidatorTests(unittest.TestCase):
    def test_complete_blueprint_passes(self) -> None:
        content = """# Assessment Blueprint
- Assessed outcome scope: LO-1
| Item ID | Outcome(s) | Intended interpretation/use | Evidence claim | Cognitive demand | Item type | Dependency | Expected time (min) | Points | Construct-irrelevant barriers | Status |
|---|---|---|---|---|---|---|---|---|---|---|
| A-1 | LO-1 | Formative feedback | Select and justify a model | Evaluate | Constructed response | independent | 12 | 10 | none identified | approved |
- Evidence level claimed: classroom-reviewed
"""
        result = run_script(
            "validate_assessment_blueprint.py",
            content,
            ".md",
            "--required-outcome",
            "LO-1",
        )
        self.assertEqual(result.returncode, 0, result.stdout + result.stderr)

    def test_comma_separated_outcomes_are_rejected_with_separator_help(self) -> None:
        content = """# Assessment Blueprint
- Assessed outcome scope: LO-1; LO-2
| Item ID | Outcome(s) | Intended interpretation/use | Evidence claim | Cognitive demand | Item type | Dependency | Expected time (min) | Points | Construct-irrelevant barriers | Status |
|---|---|---|---|---|---|---|---|---|---|---|
| A-1 | LO-1, LO-2 | Formative feedback | Explain a relationship | Analyze | Constructed response | independent | 12 | 10 | none identified | approved |
- Evidence level claimed: classroom-reviewed
"""
        result = run_script("validate_assessment_blueprint.py", content)
        self.assertEqual(result.returncode, 2, result.stdout + result.stderr)
        self.assertIn("use semicolons between identifiers", result.stdout)

    def test_blueprint_detects_missing_coverage_and_overclaim(self) -> None:
        content = """# Assessment Blueprint
- Assessed outcome scope: LO-1; LO-2
| Item ID | Outcome(s) | Intended interpretation/use | Evidence claim | Cognitive demand | Item type | Dependency | Expected time (min) | Points | Construct-irrelevant barriers | Status |
|---|---|---|---|---|---|---|---|---|---|---|
| A-1 | LO-1 | Final grade | Correct answer | Apply | Problem | independent | 10 | 10 | none identified | review |
- Evidence level claimed: formally validated
"""
        result = run_script(
            "validate_assessment_blueprint.py",
            content,
            ".md",
            "--required-outcome",
            "LO-2",
        )
        self.assertEqual(result.returncode, 2, result.stdout + result.stderr)
        self.assertIn("Required outcome is not sampled: LO-2", result.stdout)
        self.assertIn("Formal-validation claim requires", result.stdout)

    def test_retired_alignment_outcome_is_not_required(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            directory = Path(temp_dir)
            alignment = directory / "alignment-map.md"
            blueprint = directory / "assessment-blueprint.md"
            alignment.write_text(
                """| Outcome ID | Observable learning outcome | Cognitive demand | Evidence of learning | Learning mechanism | Learning activity/support | Feedback or assessment | Status |
|---|---|---|---|---|---|---|---|
| LO-1 | Analyze evidence | analyze | Analysis | deliberate practice | Cases | Rubric | approved |
| LO-2 | Defend a design | evaluate | Defense | feedback | Studio | Checklist | retired |
""",
                encoding="utf-8",
            )
            blueprint.write_text(
                """- Assessed outcome scope: LO-1
- Evidence level claimed: expert-reviewed
| Item ID | Outcome(s) | Intended interpretation/use | Evidence claim | Cognitive demand | Item type | Dependency | Expected time (min) | Points | Construct-irrelevant barriers | Status |
|---|---|---|---|---|---|---|---|---|---|---|
| A-1 | LO-1 | Judge analysis | Applies criteria | analyze | essay | independent | 20 | 10 | writing load | approved |
""",
                encoding="utf-8",
            )
            result = run_path(
                "validate_assessment_blueprint.py",
                blueprint,
                "--alignment-map",
                str(alignment),
            )
        self.assertEqual(result.returncode, 0, result.stdout + result.stderr)

    def test_retired_item_does_not_satisfy_required_outcome(self) -> None:
        content = """- Assessed outcome scope: LO-1
- Evidence level claimed: expert-reviewed
| Item ID | Outcome(s) | Intended interpretation/use | Evidence claim | Cognitive demand | Item type | Dependency | Expected time (min) | Points | Construct-irrelevant barriers | Status |
|---|---|---|---|---|---|---|---|---|---|---|
| OLD-1 | LO-1 | Historical only | Student analyzes | analyze | essay | independent | 20 | 10 | none | retired |
"""
        result = run_script(
            "validate_assessment_blueprint.py",
            content,
            ".md",
            "--required-outcome",
            "LO-1",
        )
        self.assertEqual(result.returncode, 2, result.stdout + result.stderr)
        self.assertIn("Required outcome is not sampled: LO-1", result.stdout)
        self.assertIn("Assessment blueprint contains no active items", result.stdout)

    def test_unknown_status_evidence_dependency_cycle_and_infinity_are_rejected(
        self,
    ) -> None:
        content = """- Assessed outcome scope: LO-1
| Item ID | Outcome(s) | Intended interpretation/use | Evidence claim | Cognitive demand | Item type | Dependency | Expected time (min) | Points | Construct-irrelevant barriers | Status |
|---|---|---|---|---|---|---|---|---|---|---|
| A-1 | LO-1 | Feedback | Claim | Apply | Response | A-2 | inf | 2 | none | done |
| A-2 | LO-1 | Feedback | Claim | Apply | Response | A-1 | 2 | 2 | none | review |
- Evidence level claimed: unquestionably-valid
"""
        result = run_script("validate_assessment_blueprint.py", content)
        self.assertEqual(result.returncode, 2, result.stdout + result.stderr)
        for expected in (
            "unknown status done",
            "expected time must be a positive number",
            "Circular item dependencies",
            "Unknown evidence level claimed",
        ):
            self.assertIn(expected, result.stdout)

    def test_item_outside_explicit_scope_is_rejected(self) -> None:
        content = """- Assessed outcome scope: LO-1
- Evidence level claimed: expert-reviewed
| Item ID | Outcome(s) | Intended interpretation/use | Evidence claim | Cognitive demand | Item type | Dependency | Expected time (min) | Points | Construct-irrelevant barriers | Status |
|---|---|---|---|---|---|---|---|---|---|---|
| A-1 | LO-2 | Feedback | Explains reasoning | understand | essay | independent | 10 | 5 | none identified | approved |
"""
        result = run_script("validate_assessment_blueprint.py", content)
        self.assertEqual(result.returncode, 2, result.stdout + result.stderr)
        self.assertIn(
            "Assessment item uses outcome outside declared scope: LO-2", result.stdout
        )
        self.assertIn("Required outcome is not sampled: LO-1", result.stdout)

    def test_missing_assessed_outcome_scope_is_rejected(self) -> None:
        content = """- Evidence level claimed: expert-reviewed
| Item ID | Outcome(s) | Intended interpretation/use | Evidence claim | Cognitive demand | Item type | Dependency | Expected time (min) | Points | Construct-irrelevant barriers | Status |
|---|---|---|---|---|---|---|---|---|---|---|
| A-1 | LO-1 | Feedback | Explains reasoning | understand | essay | independent | 10 | 5 | none identified | approved |
"""
        result = run_script("validate_assessment_blueprint.py", content)
        self.assertEqual(result.returncode, 2, result.stdout + result.stderr)
        self.assertIn("Missing assessed outcome scope", result.stdout)

    def test_scope_must_reference_an_active_alignment_outcome(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            directory = Path(temp_dir)
            alignment = directory / "alignment-map.md"
            blueprint = directory / "assessment-blueprint.md"
            alignment.write_text(
                """| Outcome ID | Observable learning outcome | Cognitive demand | Evidence of learning | Learning mechanism | Learning activity/support | Feedback or assessment | Status |
|---|---|---|---|---|---|---|---|
| LO-1 | Explain | understand | Explanation | retrieval | Activity | Feedback | approved |
| LO-2 | Defend | evaluate | Defense | feedback | Studio | Rubric | retired |
""",
                encoding="utf-8",
            )
            blueprint.write_text(
                """- Assessed outcome scope: LO-2
- Evidence level claimed: expert-reviewed
| Item ID | Outcome(s) | Intended interpretation/use | Evidence claim | Cognitive demand | Item type | Dependency | Expected time (min) | Points | Construct-irrelevant barriers | Status |
|---|---|---|---|---|---|---|---|---|---|---|
| A-1 | LO-2 | Feedback | Defends design | evaluate | oral defense | independent | 10 | 5 | none identified | approved |
""",
                encoding="utf-8",
            )
            result = run_path(
                "validate_assessment_blueprint.py",
                blueprint,
                "--alignment-map",
                str(alignment),
            )
        self.assertEqual(result.returncode, 2, result.stdout + result.stderr)
        self.assertIn(
            "Assessed outcome scope is not active in the alignment map: LO-2",
            result.stdout,
        )

    def test_all_active_scope_rejects_an_alignment_with_no_active_outcomes(
        self,
    ) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            directory = Path(temp_dir)
            alignment = directory / "alignment-map.md"
            blueprint = directory / "assessment-blueprint.md"
            alignment.write_text(
                """| Outcome ID | Observable learning outcome | Cognitive demand | Evidence of learning | Learning mechanism | Learning activity/support | Feedback or assessment | Status |
|---|---|---|---|---|---|---|---|
| LO-1 | Explain | understand | Explanation | retrieval | Archived activity | Archived feedback | retired |
""",
                encoding="utf-8",
            )
            blueprint.write_text(
                """- Assessed outcome scope: all-active
- Evidence level claimed: expert-reviewed
| Item ID | Outcome(s) | Intended interpretation/use | Evidence claim | Cognitive demand | Item type | Dependency | Expected time (min) | Points | Construct-irrelevant barriers | Status |
|---|---|---|---|---|---|---|---|---|---|---|
| A-1 | LO-1 | Feedback | Explains reasoning | understand | essay | independent | 10 | 5 | none identified | approved |
""",
                encoding="utf-8",
            )
            result = run_path(
                "validate_assessment_blueprint.py",
                blueprint,
                "--alignment-map",
                str(alignment),
            )
        self.assertEqual(result.returncode, 2, result.stdout + result.stderr)
        self.assertIn(
            "Assessed outcome scope all-active contains no active aligned outcomes",
            result.stdout,
        )
        self.assertIn(
            "Assessment item uses outcome outside declared scope: LO-1", result.stdout
        )

    def _demand_pair(
        self, directory: Path, outcome_demand: str, item_demands: tuple[str, ...]
    ) -> Path:
        alignment = directory / "alignment-map.md"
        blueprint = directory / "assessment-blueprint.md"
        alignment.write_text(
            "| Outcome ID | Observable learning outcome | Cognitive demand | Evidence of learning "
            "| Learning mechanism | Learning activity/support | Feedback or assessment | Status |\n"
            "|---|---|---|---|---|---|---|---|\n"
            f"| LO-1 | Decide under uncertainty | {outcome_demand} | Recommendation | contrasting cases "
            "| Comparison | Rubric | approved |\n",
            encoding="utf-8",
        )
        rows = "\n".join(
            f"| A-{index} | LO-1 | Use | Claim | {demand} | essay | independent | 10 | 5 | none identified | approved |"
            for index, demand in enumerate(item_demands, start=1)
        )
        blueprint.write_text(
            "- Assessed outcome scope: LO-1\n- Evidence level claimed: classroom-reviewed\n\n"
            "| Item ID | Outcome(s) | Intended interpretation/use | Evidence claim | Cognitive demand "
            "| Item type | Dependency | Expected time (min) | Points | Construct-irrelevant barriers | Status |\n"
            "|---|---|---|---|---|---|---|---|---|---|---|\n" + rows + "\n",
            encoding="utf-8",
        )
        return blueprint

    def test_assessment_below_outcome_demand_is_reported(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            directory = Path(temp_dir)
            blueprint = self._demand_pair(directory, "evaluate", ("apply",))
            result = run_path(
                "validate_assessment_blueprint.py",
                blueprint,
                "--alignment-map",
                str(directory / "alignment-map.md"),
            )
        self.assertEqual(result.returncode, 2, result.stdout + result.stderr)
        self.assertIn(
            "LO-1: highest active item demand apply is below the aligned outcome demand evaluate",
            result.stdout,
        )

    def test_scaffolding_item_below_outcome_demand_is_not_reported(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            directory = Path(temp_dir)
            blueprint = self._demand_pair(directory, "evaluate", ("apply", "evaluate"))
            result = run_path(
                "validate_assessment_blueprint.py",
                blueprint,
                "--alignment-map",
                str(directory / "alignment-map.md"),
            )
        self.assertEqual(result.returncode, 0, result.stdout + result.stderr)

    def test_demand_match_is_skipped_without_an_alignment_map(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            directory = Path(temp_dir)
            blueprint = self._demand_pair(directory, "evaluate", ("apply",))
            result = run_path("validate_assessment_blueprint.py", blueprint)
        self.assertEqual(result.returncode, 0, result.stdout + result.stderr)

    def test_unknown_item_cognitive_demand_is_reported(self) -> None:
        content = """- Assessed outcome scope: LO-1
- Evidence level claimed: classroom-reviewed

| Item ID | Outcome(s) | Intended interpretation/use | Evidence claim | Cognitive demand | Item type | Dependency | Expected time (min) | Points | Construct-irrelevant barriers | Status |
|---|---|---|---|---|---|---|---|---|---|---|
| A-1 | LO-1 | Feedback | Explanation | synthesize | essay | independent | 10 | 5 | none identified | approved |
"""
        result = run_script("validate_assessment_blueprint.py", content)
        self.assertEqual(result.returncode, 2, result.stdout + result.stderr)
        self.assertIn("unknown cognitive demand synthesize", result.stdout)

    def test_zero_point_formative_item_is_allowed(self) -> None:
        content = """# Assessment Blueprint
- Assessed outcome scope: LO-1
- Evidence level claimed: classroom-reviewed
| Item ID | Outcome(s) | Intended interpretation/use | Evidence claim | Cognitive demand | Item type | Dependency | Expected time (min) | Points | Construct-irrelevant barriers | Status |
|---|---|---|---|---|---|---|---|---|---|---|
| A-1 | LO-1 | Formative feedback | Explanation | understand | response | independent | 5 | 0 | none identified | approved |
"""
        result = run_script("validate_assessment_blueprint.py", content)
        self.assertEqual(result.returncode, 0, result.stdout + result.stderr)

    def test_separator_only_item_id_is_reported_without_a_traceback(self) -> None:
        content = """# Assessment Blueprint
- Assessed outcome scope: LO-1
- Evidence level claimed: classroom-reviewed
| Item ID | Outcome(s) | Intended interpretation/use | Evidence claim | Cognitive demand | Item type | Dependency | Expected time (min) | Points | Construct-irrelevant barriers | Status |
|---|---|---|---|---|---|---|---|---|---|---|
| ; | LO-1 | Feedback | Explanation | understand | response | independent | 5 | 0 | none identified | approved |
"""
        result = run_script("validate_assessment_blueprint.py", content)
        self.assertEqual(result.returncode, 2, result.stdout + result.stderr)
        self.assertIn("identifier list contains no identifiers", result.stdout)
        self.assertNotIn("Traceback", result.stderr)

    def test_duplicate_markdown_assessment_metadata_is_structural_error(self) -> None:
        base = """- Assessed outcome scope: LO-1
- Evidence level claimed: expert-reviewed
| Item ID | Outcome(s) | Intended interpretation/use | Evidence claim | Cognitive demand | Item type | Dependency | Expected time (min) | Points | Construct-irrelevant barriers | Status |
|---|---|---|---|---|---|---|---|---|---|---|
| A-1 | LO-1 | Feedback | Explanation | understand | essay | independent | 10 | 5 | none identified | approved |
"""
        cases = (
            (
                base.replace(
                    "- Evidence level claimed: expert-reviewed",
                    "- Evidence level claimed: expert-reviewed\n"
                    "- Evidence level claimed: formally-validated",
                ),
                "duplicate evidence-level claims",
            ),
            (
                base.replace(
                    "- Assessed outcome scope: LO-1",
                    "- Assessed outcome scope: LO-1\n" "- Assessed outcome scope: LO-2",
                ),
                "duplicate assessed-outcome scopes",
            ),
        )
        for content, expected in cases:
            with self.subTest(expected=expected):
                result = run_script("validate_assessment_blueprint.py", content)
                self.assertEqual(result.returncode, 1, result.stdout + result.stderr)
                self.assertIn(expected, result.stdout)

    def test_unknown_alignment_status_does_not_define_active_scope(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            directory = Path(temp_dir)
            alignment = directory / "alignment-map.md"
            blueprint = directory / "assessment-blueprint.md"
            alignment.write_text(
                """| Outcome ID | Observable learning outcome | Cognitive demand | Evidence of learning | Learning mechanism | Learning activity/support | Feedback or assessment | Status |
|---|---|---|---|---|---|---|---|
| LO-1 | Explain | understand | Explanation | retrieval | Activity | Feedback | done |
""",
                encoding="utf-8",
            )
            blueprint.write_text(
                """- Assessed outcome scope: LO-1
- Evidence level claimed: expert-reviewed
| Item ID | Outcome(s) | Intended interpretation/use | Evidence claim | Cognitive demand | Item type | Dependency | Expected time (min) | Points | Construct-irrelevant barriers | Status |
|---|---|---|---|---|---|---|---|---|---|---|
| A-1 | LO-1 | Feedback | Explanation | understand | essay | independent | 10 | 5 | none identified | approved |
""",
                encoding="utf-8",
            )
            result = run_path(
                "validate_assessment_blueprint.py",
                blueprint,
                "--alignment-map",
                str(alignment),
            )
        self.assertEqual(result.returncode, 2, result.stdout + result.stderr)
        self.assertIn(
            "Assessed outcome scope is not active in the alignment map: LO-1",
            result.stdout,
        )


class CourseCurriculumMapValidatorTests(unittest.TestCase):
    def test_coherent_course_map_passes(self) -> None:
        content = """# Course Curriculum Map
| Sequence | Module/week | Outcome ID | Developmental stage | Outcome prerequisites | Learning experience/evidence | Feedback/assessment | Expected student workload (hours) | Status |
|---|---|---|---|---|---|---|---|---|
| 1 | 1 | LO-1 | introduce | external: prerequisite course | Prediction and model sketch | Diagnostic feedback | 2 | approved |
| 2 | 2 | LO-1 | practice | none | Contrasting cases | Peer and instructor feedback | 3 | approved |
| 3 | 3 | LO-1 | assess | none | Novel transfer problem | Scored rubric | 2 | approved |
"""
        result = run_script(
            "validate_course_curriculum_map.py",
            content,
            ".md",
            "--max-hours-per-module",
            "8",
        )
        self.assertEqual(result.returncode, 0, result.stdout + result.stderr)

    MASSED_PRACTICE_MAP = """# Course Curriculum Map
| Sequence | Module/week | Outcome ID | Developmental stage | Outcome prerequisites | Learning experience/evidence | Feedback/assessment | Expected student workload (hours) | Status |
|---|---|---|---|---|---|---|---|---|
| 1 | 2 | LO-1 | introduce | none | Prediction | Diagnostic feedback | 2 | approved |
| 2 | 2 | LO-1 | practice | none | Problem set | Peer feedback | 2 | approved |
| 3 | 2 | LO-1 | practice | none | Problem set | Instructor feedback | 2 | approved |
| 4 | 6 | LO-1 | assess | none | Novel transfer problem | Scored rubric | 2 | approved |
"""

    def test_massed_practice_is_silent_without_the_flag(self) -> None:
        result = run_script(
            "validate_course_curriculum_map.py", self.MASSED_PRACTICE_MAP
        )
        self.assertEqual(result.returncode, 0, result.stdout + result.stderr)

    def test_massed_practice_is_reported_with_the_flag(self) -> None:
        result = run_script(
            "validate_course_curriculum_map.py",
            self.MASSED_PRACTICE_MAP,
            ".md",
            "--check-practice-distribution",
        )
        self.assertEqual(result.returncode, 2, result.stdout + result.stderr)
        self.assertIn("practice occurrences are massed in module/week 2", result.stdout)

    def test_distributed_practice_passes_the_flag(self) -> None:
        content = self.MASSED_PRACTICE_MAP.replace(
            "| 3 | 2 | LO-1 | practice", "| 3 | 4 | LO-1 | practice"
        )
        result = run_script(
            "validate_course_curriculum_map.py",
            content,
            ".md",
            "--check-practice-distribution",
        )
        self.assertEqual(result.returncode, 0, result.stdout + result.stderr)

    def test_blank_module_does_not_produce_a_massed_practice_claim(self) -> None:
        content = """# Course Curriculum Map
| Sequence | Module/week | Outcome ID | Developmental stage | Outcome prerequisites | Learning experience/evidence | Feedback/assessment | Expected student workload (hours) | Status |
|---|---|---|---|---|---|---|---|---|
| 1 | 1 | LO-1 | introduce | none | Prediction | Diagnostic | 1 | approved |
| 2 |  | LO-1 | practice | none | Problem set | Peer feedback | 1 | approved |
| 3 |  | LO-1 | practice | none | Problem set | Instructor feedback | 1 | approved |
"""
        result = run_script(
            "validate_course_curriculum_map.py",
            content,
            ".md",
            "--check-practice-distribution",
        )
        self.assertEqual(result.returncode, 2, result.stdout + result.stderr)
        self.assertIn("missing module/week", result.stdout)
        self.assertNotIn("massed", result.stdout)

    def test_single_practice_row_is_not_flagged_as_massed(self) -> None:
        content = """# Course Curriculum Map
| Sequence | Module/week | Outcome ID | Developmental stage | Outcome prerequisites | Learning experience/evidence | Feedback/assessment | Expected student workload (hours) | Status |
|---|---|---|---|---|---|---|---|---|
| 1 | 1 | LO-1 | introduce | none | Prediction | Diagnostic feedback | 2 | approved |
| 2 | 2 | LO-1 | practice | none | Problem set | Peer feedback | 2 | approved |
| 3 | 3 | LO-1 | assess | none | Transfer problem | Scored rubric | 2 | approved |
"""
        result = run_script(
            "validate_course_curriculum_map.py",
            content,
            ".md",
            "--check-practice-distribution",
        )
        self.assertEqual(result.returncode, 0, result.stdout + result.stderr)

    def test_course_map_detects_sequence_cycle_and_overload(self) -> None:
        content = """# Course Curriculum Map
| Sequence | Module/week | Outcome ID | Developmental stage | Outcome prerequisites | Learning experience/evidence | Feedback/assessment | Expected student workload (hours) | Status |
|---|---|---|---|---|---|---|---|---|
| 1 | 1 | LO-1 | assess | LO-2 | Final design | Scored rubric | 6 | review |
| 1 | 1 | LO-2 | introduce | LO-1 | Initial model | Diagnostic feedback | 5 | review |
"""
        result = run_script(
            "validate_course_curriculum_map.py",
            content,
            ".md",
            "--max-hours-per-module",
            "8",
        )
        self.assertEqual(result.returncode, 2, result.stdout + result.stderr)
        self.assertIn("before introduction or a declared external prior", result.stdout)
        self.assertIn("mastery/assessment appears before practice", result.stdout)
        self.assertIn("Circular outcome prerequisites", result.stdout)
        self.assertIn("exceeds limit 8", result.stdout)

    def test_sequence_is_evaluated_independently_of_physical_row_order(self) -> None:
        content = """| Sequence | Module/week | Outcome ID | Developmental stage | Outcome prerequisites | Learning experience/evidence | Feedback/assessment | Expected student workload (hours) | Status |
|---|---|---|---|---|---|---|---|---|
| 3 | 3 | LO-1 | assess | none | Transfer task | Rubric | 2 | approved |
| 1 | 1 | LO-1 | introduce | none | Initial model | Feedback | 2 | approved |
| 2 | 2 | LO-1 | practice | none | Cases | Feedback | 2 | approved |
"""
        result = run_script("validate_course_curriculum_map.py", content)
        self.assertEqual(result.returncode, 0, result.stdout + result.stderr)

    def test_same_sequence_does_not_create_false_prior_development(self) -> None:
        content = """| Sequence | Module/week | Outcome ID | Developmental stage | Outcome prerequisites | Learning experience/evidence | Feedback/assessment | Expected student workload (hours) | Status |
|---|---|---|---|---|---|---|---|---|
| 1 | 1 | LO-1 | introduce | none | Initial model | Feedback | 2 | review |
| 1 | 1 | LO-1 | assess | none | Transfer task | Rubric | 2 | review |
"""
        result = run_script("validate_course_curriculum_map.py", content)
        self.assertEqual(result.returncode, 2, result.stdout + result.stderr)
        self.assertIn("before introduction or a declared external prior", result.stdout)
        self.assertIn("mastery/assessment appears before practice", result.stdout)

    def test_practice_after_assessment_does_not_satisfy_progression(self) -> None:
        content = """| Sequence | Module/week | Outcome ID | Developmental stage | Outcome prerequisites | Learning experience/evidence | Feedback/assessment | Expected student workload (hours) | Status |
|---|---|---|---|---|---|---|---|---|
| 1 | 1 | LO-1 | introduce | none | Initial model | Feedback | 2 | approved |
| 2 | 2 | LO-1 | assess | none | Transfer task | Rubric | 2 | approved |
| 3 | 3 | LO-1 | practice | none | Cases | Feedback | 2 | approved |
"""
        result = run_script(
            "validate_course_curriculum_map.py",
            content,
            ".md",
            "--require-complete-progression",
        )
        self.assertEqual(result.returncode, 2, result.stdout + result.stderr)
        self.assertIn("mastery/assessment appears before practice", result.stdout)

    def test_mixed_external_and_internal_prerequisites_are_checked(self) -> None:
        content = """| Sequence | Module/week | Outcome ID | Developmental stage | Outcome prerequisites | Learning experience/evidence | Feedback/assessment | Expected student workload (hours) | Status |
|---|---|---|---|---|---|---|---|---|
| 1 | 1 | LO-1 | introduce | external: calculus; LO-9 | Initial model | Feedback | 2 | review |
"""
        result = run_script("validate_course_curriculum_map.py", content)
        self.assertEqual(result.returncode, 2, result.stdout + result.stderr)
        self.assertIn("unknown prerequisite outcome lo-9", result.stdout)

    def test_comma_separated_prerequisites_are_rejected_with_separator_help(
        self,
    ) -> None:
        content = """| Sequence | Module/week | Outcome ID | Developmental stage | Outcome prerequisites | Learning experience/evidence | Feedback/assessment | Expected student workload (hours) | Status |
|---|---|---|---|---|---|---|---|---|
| 1 | 1 | LO-3 | introduce | LO-1, LO-2 | Initial model | Feedback | 2 | review |
"""
        result = run_script("validate_course_curriculum_map.py", content)
        self.assertEqual(result.returncode, 2, result.stdout + result.stderr)
        self.assertIn("use semicolons between identifiers", result.stdout)

    def test_external_prerequisite_description_preserves_a_comma(self) -> None:
        content = """| Sequence | Module/week | Outcome ID | Developmental stage | Outcome prerequisites | Learning experience/evidence | Feedback/assessment | Expected student workload (hours) | Status |
|---|---|---|---|---|---|---|---|---|
| 1 | 1 | LO-1 | introduce | external: OSHA standard, 2024 edition | Initial model | Feedback | 2 | approved |
"""
        result = run_script("validate_course_curriculum_map.py", content)
        self.assertEqual(result.returncode, 0, result.stdout + result.stderr)

    def test_prerequisite_must_be_developed_before_dependent_outcome(self) -> None:
        content = """| Sequence | Module/week | Outcome ID | Developmental stage | Outcome prerequisites | Learning experience/evidence | Feedback/assessment | Expected student workload (hours) | Status |
|---|---|---|---|---|---|---|---|---|
| 1 | 1 | LO-2 | introduce | LO-1 | Dependent activity | Feedback | 2 | approved |
| 2 | 2 | LO-1 | introduce | none | Prerequisite activity | Feedback | 2 | approved |
"""
        result = run_script("validate_course_curriculum_map.py", content)
        self.assertEqual(result.returncode, 2, result.stdout + result.stderr)
        self.assertIn(
            "prerequisite outcome lo-1 has no earlier development or declared external prior",
            result.stdout,
        )

    def test_earlier_prerequisite_development_passes(self) -> None:
        content = """| Sequence | Module/week | Outcome ID | Developmental stage | Outcome prerequisites | Learning experience/evidence | Feedback/assessment | Expected student workload (hours) | Status |
|---|---|---|---|---|---|---|---|---|
| 1 | 1 | LO-1 | introduce | none | Prerequisite activity | Feedback | 2 | approved |
| 2 | 2 | LO-2 | introduce | LO-1 | Dependent activity | Feedback | 2 | approved |
"""
        result = run_script("validate_course_curriculum_map.py", content)
        self.assertEqual(result.returncode, 0, result.stdout + result.stderr)

    def test_separator_only_curriculum_outcome_is_reported_without_a_traceback(
        self,
    ) -> None:
        content = """| Sequence | Module/week | Outcome ID | Developmental stage | Outcome prerequisites | Learning experience/evidence | Feedback/assessment | Expected student workload (hours) | Status |
|---|---|---|---|---|---|---|---|---|
| 1 | 1 | ; | introduce | none | Activity | Feedback | 2 | approved |
"""
        result = run_script("validate_course_curriculum_map.py", content)
        self.assertEqual(result.returncode, 2, result.stdout + result.stderr)
        self.assertIn("identifier list contains no identifiers", result.stdout)
        self.assertNotIn("Traceback", result.stderr)

    def test_handoff_progression_requires_practice_and_terminal_evidence(self) -> None:
        content = """| Sequence | Module/week | Outcome ID | Developmental stage | Outcome prerequisites | Learning experience/evidence | Feedback/assessment | Expected student workload (hours) | Status |
|---|---|---|---|---|---|---|---|---|
| 1 | Week 1 | LO-1 | introduce | external: entry knowledge | Guided example | Minute paper | 1 | approved |
"""
        result = run_script(
            "validate_course_curriculum_map.py",
            content,
            ".md",
            "--require-complete-progression",
        )
        self.assertEqual(result.returncode, 2, result.stdout + result.stderr)
        self.assertIn("complete progression is missing practice", result.stdout)
        self.assertIn(
            "complete progression is missing mastery or assessment", result.stdout
        )

    def test_retired_rows_do_not_count_toward_current_workload(self) -> None:
        content = """| Sequence | Module/week | Outcome ID | Developmental stage | Outcome prerequisites | Learning experience/evidence | Feedback/assessment | Expected student workload (hours) | Status |
|---|---|---|---|---|---|---|---|---|
| 1 | Old module | LO-1 | introduce | external: prior | Archived activity | Archived feedback | 99 | retired |
| 2 | Current module | LO-2 | introduce | external: prior | Current activity | Feedback | 2 | approved |
"""
        result = run_script(
            "validate_course_curriculum_map.py",
            content,
            ".md",
            "--max-hours-per-module",
            "8",
        )
        self.assertEqual(result.returncode, 0, result.stdout + result.stderr)

    def test_retired_rows_do_not_shift_reported_row_numbers(self) -> None:
        content = """| Sequence | Module/week | Outcome ID | Developmental stage | Outcome prerequisites | Learning experience/evidence | Feedback/assessment | Expected student workload (hours) | Status |
|---|---|---|---|---|---|---|---|---|
| 1 | Old module | LO-1 | introduce | external: prior | Archived activity | Archived feedback | 99 | retired |
| 2 | Current module | LO-2 | introduce | external: prior | Current activity | Feedback | invalid | approved |
"""
        result = run_script("validate_course_curriculum_map.py", content)
        self.assertEqual(result.returncode, 2, result.stdout + result.stderr)
        self.assertIn("Row 2 (LO-2): workload must be a positive number", result.stdout)

    def test_retired_only_map_reports_no_active_rows(self) -> None:
        content = """| Sequence | Module/week | Outcome ID | Developmental stage | Outcome prerequisites | Learning experience/evidence | Feedback/assessment | Expected student workload (hours) | Status |
|---|---|---|---|---|---|---|---|---|
| 1 | Old module | LO-1 | introduce | external: prior | Archived activity | Archived feedback | 2 | retired |
"""
        result = run_script("validate_course_curriculum_map.py", content)
        self.assertEqual(result.returncode, 2, result.stdout + result.stderr)
        self.assertIn("Course curriculum map contains no active rows", result.stdout)

    def test_malformed_retired_row_is_still_validated_structurally(self) -> None:
        content = """| Sequence | Module/week | Outcome ID | Developmental stage | Outcome prerequisites | Learning experience/evidence | Feedback/assessment | Expected student workload (hours) | Status |
|---|---|---|---|---|---|---|---|---|
|  |  |  |  |  |  |  |  | retired |
| 1 | Week 1 | LO-1 | introduce | external: prior | Activity | Feedback | 1 | approved |
"""
        result = run_script("validate_course_curriculum_map.py", content)
        self.assertEqual(result.returncode, 2, result.stdout + result.stderr)
        self.assertIn("Row 1: missing outcome ID", result.stdout)
        self.assertIn("missing sequence", result.stdout)


class ValidatorFixtureAndCliTests(unittest.TestCase):
    CASES = (
        ("validate_design_state.py", "design_state", (), 0, 2, 1, 0),
        ("validate_alignment_map.py", "alignment", (), 0, 2, 1, 0),
        ("validate_artifact_manifest.py", "artifact_manifest", (), 0, 2, 1, 0),
        (
            "validate_assessment_blueprint.py",
            "assessment_blueprint",
            ("--required-outcome", "LO-9"),
            0,
            2,
            1,
            0,
        ),
        (
            "validate_course_curriculum_map.py",
            "course_curriculum_map",
            ("--max-hours-per-module", "8"),
            0,
            2,
            1,
            0,
        ),
    )

    def test_valid_invalid_malformed_and_boundary_fixtures(self) -> None:
        for (
            script,
            fixture_dir,
            invalid_args,
            valid_code,
            invalid_code,
            malformed_code,
            boundary_code,
        ) in self.CASES:
            with self.subTest(script=script, fixture="valid"):
                result = run_path(script, FIXTURES / fixture_dir / "valid.md")
                self.assertEqual(
                    result.returncode, valid_code, result.stdout + result.stderr
                )
            with self.subTest(script=script, fixture="invalid"):
                result = run_path(
                    script, FIXTURES / fixture_dir / "invalid.md", *invalid_args
                )
                self.assertEqual(
                    result.returncode, invalid_code, result.stdout + result.stderr
                )
            with self.subTest(script=script, fixture="malformed"):
                result = run_path(script, FIXTURES / fixture_dir / "malformed.md")
                self.assertEqual(
                    result.returncode, malformed_code, result.stdout + result.stderr
                )
            boundary_path = next((FIXTURES / fixture_dir).glob("boundary.*"))
            with self.subTest(script=script, fixture="boundary"):
                result = run_path(script, boundary_path)
                self.assertEqual(
                    result.returncode, boundary_code, result.stdout + result.stderr
                )

    def test_missing_file_returns_structural_error_for_every_validator(self) -> None:
        missing = FIXTURES / "does-not-exist.md"
        for script, *_ in self.CASES:
            with self.subTest(script=script):
                result = run_path(script, missing)
                self.assertEqual(result.returncode, 1, result.stdout + result.stderr)
                self.assertIn("ERROR:", result.stdout)

    def test_help_includes_usage_and_examples_for_every_validator(self) -> None:
        for script, *_ in self.CASES:
            with self.subTest(script=script):
                result = run_help(script)
                self.assertEqual(result.returncode, 0, result.stdout + result.stderr)
                self.assertIn("usage:", result.stdout.lower())
                self.assertIn("Example", result.stdout)

    def test_assessment_fixture_exercises_quality_rules(self) -> None:
        result = run_path(
            "validate_assessment_blueprint.py",
            FIXTURES / "assessment_blueprint" / "invalid.md",
            "--required-outcome",
            "LO-9",
        )
        self.assertEqual(result.returncode, 2, result.stdout + result.stderr)
        for expected in (
            "duplicate item ID A-1",
            "expected time must be a positive number",
            "points must be a nonnegative finite number",
            "assessment content has no item ID",
            "Required outcome is not sampled: LO-9",
            "Formal-validation claim requires",
        ):
            self.assertIn(expected, result.stdout)

    def test_assessment_empty_row_rule(self) -> None:
        content = """| Item ID | Outcome(s) | Intended interpretation/use | Evidence claim | Cognitive demand | Item type | Dependency | Expected time (min) | Points | Construct-irrelevant barriers | Status |
|---|---|---|---|---|---|---|---|---|---|---|
|  |  |  |  |  |  |  |  |  |  |  |
"""
        result = run_script("validate_assessment_blueprint.py", content)
        self.assertEqual(result.returncode, 2, result.stdout + result.stderr)
        self.assertIn("empty row", result.stdout)

    def test_course_map_fixture_exercises_coherence_rules(self) -> None:
        result = run_path(
            "validate_course_curriculum_map.py",
            FIXTURES / "course_curriculum_map" / "invalid.md",
            "--max-hours-per-module",
            "8",
        )
        self.assertEqual(result.returncode, 2, result.stdout + result.stderr)
        for expected in (
            "before introduction or a declared external prior",
            "mastery/assessment appears before practice",
            "missing feedback/assessment",
            "invalid stage launch",
            "unknown prerequisite outcome lo-9",
            "Circular outcome prerequisites",
            "workload must be a positive number",
            "exceeds limit 8",
        ):
            self.assertIn(expected, result.stdout)

    def test_course_map_missing_outcome_id_rule(self) -> None:
        content = """| Sequence | Module/week | Outcome ID | Developmental stage | Outcome prerequisites | Learning experience/evidence | Feedback/assessment | Expected student workload (hours) | Status |
|---|---|---|---|---|---|---|---|---|
| 1 | 1 |  | introduce | none | Activity | Feedback | 1 | draft |
"""
        result = run_script("validate_course_curriculum_map.py", content)
        self.assertEqual(result.returncode, 2, result.stdout + result.stderr)
        self.assertIn("missing outcome ID", result.stdout)

    def test_table_validators_report_empty_data_sets(self) -> None:
        empty_cases = (
            (
                "validate_alignment_map.py",
                "| Outcome ID | Observable learning outcome | Cognitive demand | Evidence of learning | Learning mechanism | Learning activity/support | Feedback or assessment | Status |\n|---|---|---|---|---|---|---|---|\n",
            ),
            (
                "validate_artifact_manifest.py",
                "| Artifact ID | Artifact type | Artifact family | Variant | Required variants | File or reference | Audience | Outcome(s) | Status | Validation completed | Blockers/open issues | Last reviewed | Production plan | Accessibility review | Safety review | Data-task-fit evidence |\n|---|---|---|---|---|---|---|---|---|---|---|---|---|---|---|---|\n",
            ),
            (
                "validate_assessment_blueprint.py",
                "| Item ID | Outcome(s) | Intended interpretation/use | Evidence claim | Cognitive demand | Item type | Dependency | Expected time (min) | Points | Construct-irrelevant barriers | Status |\n|---|---|---|---|---|---|---|---|---|---|---|\n",
            ),
            (
                "validate_course_curriculum_map.py",
                "| Sequence | Module/week | Outcome ID | Developmental stage | Outcome prerequisites | Learning experience/evidence | Feedback/assessment | Expected student workload (hours) | Status |\n|---|---|---|---|---|---|---|---|---|\n",
            ),
        )
        for script, content in empty_cases:
            with self.subTest(script=script):
                result = run_script(script, content)
                self.assertEqual(result.returncode, 2, result.stdout + result.stderr)


class ProjectValidatorTests(unittest.TestCase):
    BRIEF = """# Course Design Brief
- Schema version: 1.0
- Engagement tier: Project
## Course context
- Course or module: BIO 101
## Intended learning
- Learning outcomes: Explain a mechanism.
## Access, participation, and belonging
- Known access constraints without sensitive student details: None identified.
## Constraints
- Technology and format: Markdown.
## Implementation load
- Minimum viable fallback: Printable activity.
## Collaboration
- Interaction level: Co-design
- Requested artifacts: Student worksheet.
## Status
### Confirmed
- Outcome approved.
### Assumed
- Prior knowledge present.
### Open
- Accessibility review pending.
### Current phase
- Produce aligned artifacts.
### Next decision
- Review the worksheet.
"""
    ALIGNMENT = """# Alignment Map
- Schema version: 1.0
| Outcome ID | Observable learning outcome | Cognitive demand | Evidence of learning | Learning mechanism | Learning activity/support | Feedback or assessment | Status |
|---|---|---|---|---|---|---|---|
| LO-1 | Explain a mechanism | understand | Explanation | contrasting cases | Comparison | Feedback | approved |
"""

    def test_minimal_indexed_project_passes(self) -> None:
        index = """# Project Index
- Schema version: 1.0
- Engagement tier: Project
| State file | Purpose | Authority/owner | Schema version | Status | Last updated | Notes |
|---|---|---|---|---|---|---|
| course-design-brief.md | design authority | course owner | 1.0 | approved | 2026-08-01 | current |
| alignment-map.md | alignment authority | course owner | 1.0 | approved | 2026-08-01 | current |
"""
        result = run_project(
            {
                "project-index.md": index,
                "course-design-brief.md": self.BRIEF,
                "alignment-map.md": self.ALIGNMENT,
            }
        )
        self.assertEqual(result.returncode, 0, result.stdout + result.stderr)

    def test_project_detects_unknown_cross_file_outcome(self) -> None:
        index = """# Project Index
- Schema version: 1.0
- Engagement tier: Project
| State file | Purpose | Authority/owner | Schema version | Status | Last updated | Notes |
|---|---|---|---|---|---|---|
| course-design-brief.md | design authority | course owner | 1.0 | approved | 2026-08-01 | current |
| alignment-map.md | alignment authority | course owner | 1.0 | approved | 2026-08-01 | current |
| artifact-manifest.md | artifact authority | course owner | 1.0 | review | 2026-08-01 | current |
"""
        manifest = """# Artifact Manifest
- Schema version: 1.0
| Artifact ID | Artifact type | Artifact family | Variant | Required variants | File or reference | Audience | Outcome(s) | Status | Validation completed | Blockers/open issues | Last reviewed | Production plan | Accessibility review | Safety review | Data-task-fit evidence |
|---|---|---|---|---|---|---|---|---|---|---|---|---|---|---|---|
| WS-1 | markdown |  |  |  | https://example.edu/ws | student | LO-9 | draft | manual | none | 2026-08-01 | https://example.edu/plan | https://example.edu/access | not required | not applicable — no dataset |
"""
        result = run_project(
            {
                "project-index.md": index,
                "course-design-brief.md": self.BRIEF,
                "alignment-map.md": self.ALIGNMENT,
                "artifact-manifest.md": manifest,
            }
        )
        self.assertEqual(result.returncode, 2, result.stdout + result.stderr)
        self.assertIn("outcome is not defined in alignment-map.md: lo-9", result.stdout)

    DATA_INDEX = """# Project Index
- Schema version: 1.0
- Engagement tier: Project
| State file | Purpose | Authority/owner | Schema version | Status | Last updated | Notes |
|---|---|---|---|---|---|---|
| course-design-brief.md | design authority | course owner | 1.0 | approved | 2026-08-01 | current |
| alignment-map.md | alignment authority | course owner | 1.0 | approved | 2026-08-01 | current |
| artifact-manifest.md | artifact authority | course owner | 1.0 | review | 2026-08-01 | current |
"""
    DATA_MANIFEST = """# Artifact Manifest
- Schema version: 1.0
| Artifact ID | Artifact type | Artifact family | Variant | Required variants | File or reference | Audience | Outcome(s) | Status | Validation completed | Blockers/open issues | Last reviewed | Production plan | Accessibility review | Safety review | Data-task-fit evidence |
|---|---|---|---|---|---|---|---|---|---|---|---|---|---|---|---|
| WS-1 | markdown |  |  |  | https://example.edu/ws | student | LO-1 | draft | manual; data-task-fit | none | 2026-08-01 | https://example.edu/plan | https://example.edu/access | not required | data-task-record.md |
"""
    DATA_RECORD = """# Data–Task Fit Record
- Schema version: 1.0
| Artifact ID | Dataset file | Dataset SHA-256 | Dataset version or date | Worksheet | Representation | Column roles | Expected student output | Intended interpretation | Execution method | Execution evidence | Executed on | Result |
|---|---|---|---|---|---|---|---|---|---|---|---|---|
| WS-1 | lab.csv | c92da196adf584e997392a15e62474fc7ca8ffc59ca1495adfc874b69bf410e6 | 2026-08-01 |  | scatter | x=mass_kg; y=extension_mm | Fitted line | Extension rises with mass | manual |  | 2026-08-01 | Produced |
"""
    LAB_CSV = "mass_kg,extension_mm\n0.5,2.1\n1.0,4.3\n1.5,6.0\n2.0,8.2\n"

    def _data_index_with_record(self) -> str:
        return self.DATA_INDEX + (
            "| data-task-record.md | data-task fit | course owner | 1.0 | approved "
            "| 2026-08-01 | current |\n"
        )

    def test_fit_claim_without_a_record_cannot_be_re_executed(self) -> None:
        """A token pointing at nothing the project indexes is unverifiable."""
        result = run_project(
            {
                "project-index.md": self.DATA_INDEX,
                "course-design-brief.md": self.BRIEF,
                "alignment-map.md": self.ALIGNMENT,
                "artifact-manifest.md": self.DATA_MANIFEST,
            }
        )
        self.assertEqual(result.returncode, 2, result.stdout + result.stderr)
        self.assertIn("no active data-task-record.md", result.stdout)

    def test_fit_claim_with_a_re_executable_record_passes(self) -> None:
        result = run_project(
            {
                "project-index.md": self._data_index_with_record(),
                "course-design-brief.md": self.BRIEF,
                "alignment-map.md": self.ALIGNMENT,
                "artifact-manifest.md": self.DATA_MANIFEST,
                "data-task-record.md": self.DATA_RECORD,
                "lab.csv": self.LAB_CSV,
            }
        )
        self.assertEqual(result.returncode, 0, result.stdout + result.stderr)

    def test_project_rechecks_the_recorded_claim(self) -> None:
        """A record that no longer matches its dataset fails at project level."""
        result = run_project(
            {
                "project-index.md": self._data_index_with_record(),
                "course-design-brief.md": self.BRIEF,
                "alignment-map.md": self.ALIGNMENT,
                "artifact-manifest.md": self.DATA_MANIFEST,
                "data-task-record.md": self.DATA_RECORD,
                "lab.csv": "mass_kg,extension_mm\nlight,2.1\nheavy,4.3\nmid,6.0\n",
            }
        )
        self.assertEqual(result.returncode, 2, result.stdout + result.stderr)
        self.assertIn("rechecking the recorded claim failed", result.stdout)

    def test_evidence_must_point_at_the_record_not_merely_share_an_id(self) -> None:
        """Matching artifact IDs is not the same as pointing at the record.

        Without this, any Markdown reference satisfies the evidence cell as
        long as some record elsewhere happens to carry the same artifact ID.
        """
        result = run_project(
            {
                "project-index.md": self._data_index_with_record(),
                "course-design-brief.md": self.BRIEF,
                "alignment-map.md": self.ALIGNMENT,
                "artifact-manifest.md": self.DATA_MANIFEST.replace(
                    "| data-task-record.md |", "| unrelated-notes.md |"
                ),
                "data-task-record.md": self.DATA_RECORD,
                "unrelated-notes.md": "# Notes\n",
                "lab.csv": self.LAB_CSV,
            }
        )
        self.assertEqual(result.returncode, 2, result.stdout + result.stderr)
        self.assertIn(
            "points to unrelated-notes.md rather than the active", result.stdout
        )

    def test_record_row_for_an_unknown_artifact_is_reported(self) -> None:
        result = run_project(
            {
                "project-index.md": self._data_index_with_record(),
                "course-design-brief.md": self.BRIEF,
                "alignment-map.md": self.ALIGNMENT,
                "artifact-manifest.md": self.DATA_MANIFEST,
                "data-task-record.md": self.DATA_RECORD.replace("| WS-1 | lab.csv", "| WS-9 | lab.csv"),
                "lab.csv": self.LAB_CSV,
            }
        )
        self.assertEqual(result.returncode, 2, result.stdout + result.stderr)
        self.assertIn("is not active in artifact-manifest.md", result.stdout)

    def test_project_requires_active_index_paths(self) -> None:
        index = """- Schema version: 1.0
- Engagement tier: Project
| State file | Purpose | Authority/owner | Schema version | Status | Last updated | Notes |
|---|---|---|---|---|---|---|
| missing.md | state | course owner | 1.0 | draft | 2026-08-01 | current |
"""
        result = run_project({"project-index.md": index})
        self.assertEqual(result.returncode, 2, result.stdout + result.stderr)
        self.assertIn("active state file does not exist", result.stdout)

    def test_project_rejects_duplicate_resolved_state_file_aliases(self) -> None:
        index = """# Project Index
- Schema version: 1.0
- Engagement tier: Project
| State file | Purpose | Authority/owner | Schema version | Status | Last updated | Notes |
|---|---|---|---|---|---|---|
| course-design-brief.md | design authority | course owner | 1.0 | approved | 2026-08-01 | current |
| ./course-design-brief.md | duplicate alias | course owner | 1.0 | approved | 2026-08-01 | duplicate |
"""
        result = run_project(
            {
                "project-index.md": index,
                "course-design-brief.md": self.BRIEF,
            },
            "--design-profile",
            "establish",
        )
        self.assertEqual(result.returncode, 2, result.stdout + result.stderr)
        self.assertIn("resolves to the same target as row 1", result.stdout)

    def test_empty_handoff_index_is_rejected(self) -> None:
        index = """# Project Index
- Schema version: 1.0
- Engagement tier: Project
| State file | Purpose | Authority/owner | Schema version | Status | Last updated | Notes |
|---|---|---|---|---|---|---|
"""
        result = run_project({"project-index.md": index}, "--design-profile", "handoff")
        self.assertEqual(result.returncode, 2, result.stdout + result.stderr)
        self.assertIn("contains no state-file rows", result.stdout)
        self.assertIn("handoff profile requires", result.stdout)

    def test_project_requires_one_development_schema(self) -> None:
        index = """# Project Index
- Schema version: 1.0
- Engagement tier: Project
| State file | Purpose | Authority/owner | Schema version | Status | Last updated | Notes |
|---|---|---|---|---|---|---|
| course-design-brief.md | design authority | course owner | 2.0 | approved | 2026-08-01 | current |
"""
        other_schema_brief = self.BRIEF.replace(
            "Schema version: 1.0", "Schema version: 2.0"
        )
        result = run_project(
            {"project-index.md": index, "course-design-brief.md": other_schema_brief},
            "--design-profile",
            "establish",
        )
        self.assertEqual(result.returncode, 2, result.stdout + result.stderr)
        self.assertIn(
            "does not match project-index.md development schema", result.stdout
        )

    def test_project_requires_matching_state_dates(self) -> None:
        index = """# Project Index
- Schema version: 1.0
- Last updated: 2026-08-01
- Engagement tier: Project
| State file | Purpose | Authority/owner | Schema version | Status | Last updated | Notes |
|---|---|---|---|---|---|---|
| course-design-brief.md | design authority | course owner | 1.0 | approved | 2026-08-01 | current |
"""
        brief = self.BRIEF.replace(
            "- Schema version: 1.0",
            "- Schema version: 1.0\n- Last updated: 2026-07-31",
        )
        result = run_project(
            {"project-index.md": index, "course-design-brief.md": brief},
            "--design-profile",
            "establish",
        )
        self.assertEqual(result.returncode, 2, result.stdout + result.stderr)
        self.assertIn(
            "last updated 2026-07-31 does not match index 2026-08-01", result.stdout
        )

    def test_project_requires_index_and_state_dates(self) -> None:
        index = """# Project Index
- Schema version: 1.0
- Last updated:
- Engagement tier: Project
| State file | Purpose | Authority/owner | Schema version | Status | Last updated | Notes |
|---|---|---|---|---|---|---|
| course-design-brief.md | design authority | course owner | 1.0 | approved | 2026-08-01 | current |
"""
        brief = self.BRIEF.replace(
            "- Schema version: 1.0",
            "- Schema version: 1.0\n- Last updated:",
        )
        result = run_project(
            {"project-index.md": index, "course-design-brief.md": brief},
            "--design-profile",
            "establish",
        )
        self.assertEqual(result.returncode, 2, result.stdout + result.stderr)
        self.assertIn("project-index.md: missing last-updated date", result.stdout)
        self.assertIn(
            "course-design-brief.md: missing last-updated date", result.stdout
        )

    def test_project_passes_alignment_to_blueprint_coverage_check(self) -> None:
        index = """# Project Index
- Schema version: 1.0
- Engagement tier: Project
| State file | Purpose | Authority/owner | Schema version | Status | Last updated | Notes |
|---|---|---|---|---|---|---|
| course-design-brief.md | design authority | course owner | 1.0 | approved | 2026-08-01 | current |
| alignment-map.md | alignment authority | course owner | 1.0 | approved | 2026-08-01 | current |
| assessment-blueprint.md | assessment authority | course owner | 1.0 | review | 2026-08-01 | current |
"""
        alignment = self.ALIGNMENT + (
            "| LO-2 | Defend a design | evaluate | Oral defense | feedback | Studio | Rubric | approved |\n"
        )
        blueprint = """# Assessment Blueprint
- Schema version: 1.0
- Assessed outcome scope: all-active
- Evidence level claimed: expert-reviewed
| Item ID | Outcome(s) | Intended interpretation/use | Evidence claim | Cognitive demand | Item type | Dependency | Expected time (min) | Points | Construct-irrelevant barriers | Status |
|---|---|---|---|---|---|---|---|---|---|---|
| A-1 | LO-1 | Judge explanation | Explains mechanism | understand | essay | independent | 20 | 10 | writing load | approved |
"""
        result = run_project(
            {
                "project-index.md": index,
                "course-design-brief.md": self.BRIEF,
                "alignment-map.md": alignment,
                "assessment-blueprint.md": blueprint,
            }
        )
        self.assertEqual(result.returncode, 2, result.stdout + result.stderr)
        self.assertIn("Required outcome is not sampled: LO-2", result.stdout)

    def test_project_respects_explicit_assessment_scope(self) -> None:
        index = """# Project Index
- Schema version: 1.0
- Engagement tier: Project
| State file | Purpose | Authority/owner | Schema version | Status | Last updated | Notes |
|---|---|---|---|---|---|---|
| course-design-brief.md | design authority | course owner | 1.0 | approved | 2026-08-01 | current |
| alignment-map.md | alignment authority | course owner | 1.0 | approved | 2026-08-01 | current |
| assessment-blueprint.md | assessment authority | course owner | 1.0 | review | 2026-08-01 | current |
"""
        alignment = self.ALIGNMENT + (
            "| LO-2 | Defend a design | evaluate | Oral defense | feedback | Studio | Rubric | approved |\n"
        )
        blueprint = """# Assessment Blueprint
- Schema version: 1.0
- Assessed outcome scope: LO-1
- Evidence level claimed: expert-reviewed
| Item ID | Outcome(s) | Intended interpretation/use | Evidence claim | Cognitive demand | Item type | Dependency | Expected time (min) | Points | Construct-irrelevant barriers | Status |
|---|---|---|---|---|---|---|---|---|---|---|
| A-1 | LO-1 | Judge explanation | Explains mechanism | understand | essay | independent | 20 | 10 | writing load | approved |
"""
        result = run_project(
            {
                "project-index.md": index,
                "course-design-brief.md": self.BRIEF,
                "alignment-map.md": alignment,
                "assessment-blueprint.md": blueprint,
            }
        )
        self.assertEqual(result.returncode, 0, result.stdout + result.stderr)

    def test_course_tier_requires_curriculum_map_during_production(self) -> None:
        index = """# Project Index
- Schema version: 1.0
- Engagement tier: Course
| State file | Purpose | Authority/owner | Schema version | Status | Last updated | Notes |
|---|---|---|---|---|---|---|
| course-design-brief.md | design authority | course owner | 1.0 | approved | 2026-08-01 | current |
| alignment-map.md | alignment authority | course owner | 1.0 | approved | 2026-08-01 | current |
"""
        result = run_project(
            {
                "project-index.md": index,
                "course-design-brief.md": self.BRIEF.replace(
                    "Engagement tier: Project", "Engagement tier: Course"
                ),
                "alignment-map.md": self.ALIGNMENT,
            }
        )
        self.assertEqual(result.returncode, 2, result.stdout + result.stderr)
        self.assertIn(
            "requires an active project-index entry for course-curriculum-map.md",
            result.stdout,
        )

    def test_course_tier_map_covers_every_active_aligned_outcome(self) -> None:
        index = """# Project Index
- Schema version: 1.0
- Engagement tier: Course
| State file | Purpose | Authority/owner | Schema version | Status | Last updated | Notes |
|---|---|---|---|---|---|---|
| course-design-brief.md | design authority | course owner | 1.0 | approved | 2026-08-01 | current |
| alignment-map.md | alignment authority | course owner | 1.0 | approved | 2026-08-01 | current |
| course-curriculum-map.md | course sequence | course owner | 1.0 | review | 2026-08-01 | current |
"""
        alignment = self.ALIGNMENT + (
            "| LO-2 | Defend a design | evaluate | Oral defense | feedback | Studio | Rubric | approved |\n"
        )
        course_map = """# Course Curriculum Map
- Schema version: 1.0
| Sequence | Module/week | Outcome ID | Developmental stage | Outcome prerequisites | Learning experience/evidence | Feedback/assessment | Expected student workload (hours) | Status |
|---|---|---|---|---|---|---|---|---|
| 1 | Week 1 | LO-1 | introduce | external: prior course | Guided model | Feedback | 2 | approved |
"""
        result = run_project(
            {
                "project-index.md": index,
                "course-design-brief.md": self.BRIEF.replace(
                    "Engagement tier: Project", "Engagement tier: Course"
                ),
                "alignment-map.md": alignment,
                "course-curriculum-map.md": course_map,
            }
        )
        self.assertEqual(result.returncode, 2, result.stdout + result.stderr)
        self.assertIn(
            "active aligned outcome is not mapped for the Course tier: lo-2",
            result.stdout,
        )

    def test_project_rejects_state_path_outside_project_root(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            root = Path(temp_dir)
            project = root / "project"
            external = root / "external"
            project.mkdir()
            external.mkdir()
            (external / "course-design-brief.md").write_text(
                self.BRIEF, encoding="utf-8"
            )
            (project / "project-index.md").write_text(
                """# Project Index
- Schema version: 1.0
- Engagement tier: Project
| State file | Purpose | Authority/owner | Schema version | Status | Last updated | Notes |
|---|---|---|---|---|---|---|
| ../external/course-design-brief.md | design authority | course owner | 1.0 | approved | 2026-08-01 | external |
""",
                encoding="utf-8",
            )
            result = subprocess.run(
                [
                    sys.executable,
                    str(SCRIPTS / "validate_project.py"),
                    str(project),
                    "--design-profile",
                    "establish",
                ],
                check=False,
                capture_output=True,
                text=True,
            )
        self.assertEqual(result.returncode, 2, result.stdout + result.stderr)
        self.assertIn(
            "state file resolves outside the project directory", result.stdout
        )

    def test_project_rejects_symlinked_index_outside_project_root(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            root = Path(temp_dir)
            project = root / "project"
            project.mkdir()
            external_index = root / "external-project-index.md"
            external_index.write_text(
                """# Project Index
- Schema version: 1.0
- Engagement tier: Project
| State file | Purpose | Authority/owner | Schema version | Status | Last updated | Notes |
|---|---|---|---|---|---|---|
""",
                encoding="utf-8",
            )
            (project / "project-index.md").symlink_to(external_index)
            result = subprocess.run(
                [
                    sys.executable,
                    str(SCRIPTS / "validate_project.py"),
                    str(project),
                ],
                check=False,
                capture_output=True,
                text=True,
            )
        self.assertEqual(result.returncode, 2, result.stdout + result.stderr)
        self.assertIn(
            "project-index.md resolves outside the project directory", result.stdout
        )

    def test_project_reports_an_unresolvable_index_symlink(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            project = Path(temp_dir)
            (project / "project-index.md").symlink_to("project-index.md")
            result = subprocess.run(
                [
                    sys.executable,
                    str(SCRIPTS / "validate_project.py"),
                    str(project),
                ],
                check=False,
                capture_output=True,
                text=True,
            )
        self.assertEqual(result.returncode, 1, result.stdout + result.stderr)
        self.assertIn("Cannot resolve project-index.md", result.stdout)
        self.assertNotIn("Traceback", result.stderr)

    def test_project_reports_an_unresolvable_state_symlink(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            project = Path(temp_dir)
            (project / "course-design-brief.md").symlink_to("course-design-brief.md")
            (project / "project-index.md").write_text(
                """# Project Index
- Schema version: 1.0
- Engagement tier: Project
| State file | Purpose | Authority/owner | Schema version | Status | Last updated | Notes |
|---|---|---|---|---|---|---|
| course-design-brief.md | design authority | course owner | 1.0 | approved | 2026-08-01 | current |
""",
                encoding="utf-8",
            )
            result = subprocess.run(
                [
                    sys.executable,
                    str(SCRIPTS / "validate_project.py"),
                    str(project),
                    "--design-profile",
                    "establish",
                ],
                check=False,
                capture_output=True,
                text=True,
            )
        self.assertEqual(result.returncode, 1, result.stdout + result.stderr)
        self.assertIn("Cannot resolve state file course-design-brief.md", result.stdout)
        self.assertNotIn("Traceback", result.stderr)

    def test_course_handoff_enforces_complete_progression(self) -> None:
        index = """# Project Index
- Schema version: 1.0
- Engagement tier: Course
| State file | Purpose | Authority/owner | Schema version | Status | Last updated | Notes |
|---|---|---|---|---|---|---|
| course-design-brief.md | design authority | course owner | 1.0 | approved | 2026-08-01 | current |
| alignment-map.md | alignment authority | course owner | 1.0 | approved | 2026-08-01 | current |
| course-curriculum-map.md | course sequence | course owner | 1.0 | review | 2026-08-01 | current |
| artifact-manifest.md | artifact authority | course owner | 1.0 | review | 2026-08-01 | current |
| design-log.md | decisions | course owner | 1.0 | approved | 2026-08-01 | current |
| source-register.md | provenance | course owner | 1.0 | approved | 2026-08-01 | current |
| capability-manifest.md | capabilities | course owner | 1.0 | approved | 2026-08-01 | current |
"""
        brief = self.BRIEF.replace(
            "Engagement tier: Project", "Engagement tier: Course"
        ).replace(
            "- Known access constraints without sensitive student details: None identified.",
            "- Known access constraints without sensitive student details: None identified.\n"
            "- Accessibility contact, review, procurement, or exception process: Accessibility office.",
        )
        course_map = """# Course Curriculum Map
- Schema version: 1.0
| Sequence | Module/week | Outcome ID | Developmental stage | Outcome prerequisites | Learning experience/evidence | Feedback/assessment | Expected student workload (hours) | Status |
|---|---|---|---|---|---|---|---|---|
| 1 | Week 1 | LO-1 | introduce | external: prior course | Guided model | Feedback | 2 | approved |
"""
        manifest = """# Artifact Manifest
- Schema version: 1.0
| Artifact ID | Artifact type | Artifact family | Variant | Required variants | File or reference | Audience | Outcome(s) | Status | Validation completed | Blockers/open issues | Last reviewed | Production plan | Accessibility review | Safety review | Data-task-fit evidence |
|---|---|---|---|---|---|---|---|---|---|---|---|---|---|---|---|
| WS-1 | markdown |  |  |  | https://example.edu/ws | student | LO-1 | draft | manual | none | 2026-08-01 | not required | pending | not required | not applicable — no dataset |
"""
        state_stub = "# State\n- Schema version: 1.0\n"
        result = run_project(
            {
                "project-index.md": index,
                "course-design-brief.md": brief,
                "alignment-map.md": self.ALIGNMENT,
                "course-curriculum-map.md": course_map,
                "artifact-manifest.md": manifest,
                "design-log.md": state_stub,
                "source-register.md": state_stub,
                "capability-manifest.md": state_stub,
            },
            "--design-profile",
            "handoff",
        )
        self.assertEqual(result.returncode, 2, result.stdout + result.stderr)
        self.assertIn("complete progression is missing practice", result.stdout)
        self.assertIn(
            "complete progression is missing mastery or assessment", result.stdout
        )

    def test_missing_project_directory_is_structural_error(self) -> None:
        result = run_path("validate_project.py", FIXTURES / "does-not-exist")
        self.assertEqual(result.returncode, 1, result.stdout + result.stderr)
        self.assertIn("Project directory does not exist", result.stdout)

    def test_project_help_includes_usage_and_examples(self) -> None:
        result = run_help("validate_project.py")
        self.assertEqual(result.returncode, 0, result.stdout + result.stderr)
        self.assertIn("usage:", result.stdout.lower())
        self.assertIn("Examples", result.stdout)


class RepositoryIntegrityTests(unittest.TestCase):
    def test_repository_links_and_package_inventory_pass(self) -> None:
        result = subprocess.run(
            [sys.executable, str(ROOT / "tests" / "check_repository.py")],
            cwd=ROOT,
            check=False,
            capture_output=True,
            text=True,
        )
        self.assertEqual(result.returncode, 0, result.stdout + result.stderr)

    def test_packaged_archive_matches_the_package(self) -> None:
        """The tracked upload zip ships whatever it last contained."""
        import zipfile

        archive = ROOT / "course-development-partner.zip"
        if not archive.is_file():
            self.skipTest("no packaged archive is tracked")
        sys.path.insert(0, str(ROOT / "tests"))
        import check_repository

        expected = check_repository.actual_inventory()
        with zipfile.ZipFile(archive) as bundle:
            packaged = {
                name: bundle.read(name)
                for name in bundle.namelist()
                if not name.endswith("/")
            }
        self.assertEqual(
            set(packaged),
            set(expected),
            "course-development-partner.zip does not match the package; rebuild it",
        )
        # Same names with different bytes ships the same defect under a green check.
        stale = [path for path in expected if (ROOT / path).read_bytes() != packaged[path]]
        self.assertEqual(
            stale, [], "course-development-partner.zip holds stale copies; rebuild it"
        )

    def _archive_issues(self, build) -> list[str]:
        sys.path.insert(0, str(ROOT / "tests"))
        import check_repository

        original = check_repository.ARCHIVE
        with tempfile.TemporaryDirectory() as temp_dir:
            path = Path(temp_dir) / "candidate.zip"
            build(check_repository, path)
            check_repository.ARCHIVE = path
            try:
                errors, issues = check_repository.check_package_archive()
            finally:
                check_repository.ARCHIVE = original
        self.assertEqual(errors, [])
        return issues

    def test_archive_missing_a_file_is_reported(self) -> None:
        """The check must fail on a bad zip, not just pass on a good one."""
        import zipfile

        def build(module, path: Path) -> None:
            with zipfile.ZipFile(path, "w") as bundle:
                for name in module.actual_inventory()[:-1]:
                    bundle.write(ROOT / name, name)

        issues = self._archive_issues(build)
        self.assertTrue(any("is missing packaged file" in item for item in issues), issues)

    def test_archive_with_stale_contents_is_reported(self) -> None:
        """Every expected name present, one file's bytes behind the source."""
        import zipfile

        def build(module, path: Path) -> None:
            names = module.actual_inventory()
            with zipfile.ZipFile(path, "w") as bundle:
                for name in names:
                    if name == names[0]:
                        bundle.writestr(name, "outdated content")
                    else:
                        bundle.write(ROOT / name, name)

        issues = self._archive_issues(build)
        self.assertTrue(any("holds a stale copy of" in item for item in issues), issues)

    def test_inventory_drift_returns_two(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            inventory = Path(temp_dir) / "inventory.txt"
            inventory.write_text(
                "course-development-partner/SKILL.md\n", encoding="utf-8"
            )
            result = subprocess.run(
                [
                    sys.executable,
                    str(ROOT / "tests" / "check_repository.py"),
                    "--inventory",
                    str(inventory),
                ],
                cwd=ROOT,
                check=False,
                capture_output=True,
                text=True,
            )
        self.assertEqual(result.returncode, 2, result.stdout + result.stderr)
        self.assertIn("unrecorded file", result.stdout)

    def test_missing_inventory_returns_one(self) -> None:
        result = subprocess.run(
            [
                sys.executable,
                str(ROOT / "tests" / "check_repository.py"),
                "--inventory",
                "tests/does-not-exist.txt",
            ],
            cwd=ROOT,
            check=False,
            capture_output=True,
            text=True,
        )
        self.assertEqual(result.returncode, 1, result.stdout + result.stderr)
        self.assertIn("ERROR:", result.stdout)

    def test_privacy_screen_passes_with_bounded_claim(self) -> None:
        result = subprocess.run(
            [sys.executable, str(ROOT / "tests" / "audit_privacy.py")],
            cwd=ROOT,
            check=False,
            capture_output=True,
            text=True,
        )
        self.assertEqual(result.returncode, 0, result.stdout + result.stderr)
        self.assertIn("human review still required", result.stdout)

    def test_forward_record_hashes_match_current_inputs(self) -> None:
        record = json.loads(
            (ROOT / "tests" / "remediation-forward-test-results.json").read_text(
                encoding="utf-8"
            )
        )
        scenario_hash = hashlib.sha256(
            (ROOT / "tests" / "faculty-review-scenarios.md").read_bytes()
        ).hexdigest()
        rubric_hash = hashlib.sha256(
            (ROOT / "tests" / "evaluator-rubric.md").read_bytes()
        ).hexdigest()
        self.assertEqual(record["current_scenario_file_sha256"], scenario_hash)
        self.assertEqual(record["current_rubric_file_sha256"], rubric_hash)
        self.assertEqual(record["evidence_status"], "historical-exploratory-only")


class DatasetValidatorTests(unittest.TestCase):
    """The reported failure: a scatter plot requested from categorical totals."""

    CATEGORICAL = "region,total\nNorth,120\nSouth,90\nEast,140\nWest,75\n"
    PAIRED = "mass_kg,extension_mm\n0.5,2.1\n1.0,4.3\n1.5,6.0\n2.0,8.2\n"
    IDS = "student_id,course_code\n1001,101\n1002,102\n1003,103\n1004,104\n"
    ORDERED = "week,yield\n1,4.2\n2,5.1\n3,6.3\n"

    def test_scatter_from_categorical_totals_is_rejected(self) -> None:
        result = run_script(
            "validate_dataset.py", self.CATEGORICAL, ".csv",
            "--representation", "scatter", "--x", "region", "--y", "total",
        )
        self.assertEqual(result.returncode, 2, result.stdout + result.stderr)
        self.assertIn("must be quantitative", result.stdout)

    def test_same_data_supports_a_bar_chart(self) -> None:
        result = run_script(
            "validate_dataset.py", self.CATEGORICAL, ".csv",
            "--representation", "bar", "--category", "region", "--value", "total",
        )
        self.assertEqual(result.returncode, 0, result.stdout + result.stderr)

    def test_paired_quantitative_supports_scatter(self) -> None:
        result = run_script(
            "validate_dataset.py", self.PAIRED, ".csv",
            "--representation", "scatter", "--x", "mass_kg", "--y", "extension_mm",
        )
        self.assertEqual(result.returncode, 0, result.stdout + result.stderr)

    def test_identifier_columns_do_not_satisfy_scatter(self) -> None:
        """Two numeric columns are not two measurements."""
        result = run_script(
            "validate_dataset.py", self.IDS, ".csv",
            "--representation", "scatter", "--x", "student_id", "--y", "course_code",
        )
        self.assertEqual(result.returncode, 2, result.stdout + result.stderr)
        self.assertIn("looks like an identifier", result.stdout)

    def test_counts_are_not_mistaken_for_identifiers(self) -> None:
        """Whole, unique, non-negative describes totals as well as IDs."""
        result = run_script(
            "validate_dataset.py",
            "trial,reading\n1,120\n2,90\n3,140\n4,75\n",
            ".csv",
            "--representation", "line", "--order", "trial", "--y", "reading",
        )
        self.assertEqual(result.returncode, 0, result.stdout + result.stderr)

    def test_line_chart_requires_an_ordered_column(self) -> None:
        result = run_script(
            "validate_dataset.py", "measurement\n4.2\n5.1\n6.3\n", ".csv",
            "--representation", "line",
        )
        self.assertEqual(result.returncode, 2, result.stdout + result.stderr)
        self.assertIn("requires explicit column roles", result.stdout)

    def test_non_ordered_column_named_as_order_is_rejected(self) -> None:
        result = run_script(
            "validate_dataset.py", "label,value\nz,4.2\na,5.1\nm,6.3\n", ".csv",
            "--representation", "line", "--order", "label", "--y", "value",
        )
        self.assertEqual(result.returncode, 2, result.stdout + result.stderr)
        self.assertIn("not an ordered or time variable", result.stdout)

    def test_ordered_column_passes(self) -> None:
        result = run_script(
            "validate_dataset.py", self.ORDERED, ".csv",
            "--representation", "line", "--order", "week", "--y", "yield",
        )
        self.assertEqual(result.returncode, 0, result.stdout + result.stderr)

    def test_named_column_that_does_not_exist_is_reported(self) -> None:
        result = run_script(
            "validate_dataset.py", self.PAIRED, ".csv",
            "--representation", "scatter", "--x", "mass_kg", "--y", "temperature",
        )
        self.assertEqual(result.returncode, 2, result.stdout + result.stderr)
        self.assertIn("does not exist in the dataset", result.stdout)

    def test_negative_min_rows_is_an_error(self) -> None:
        result = run_script(
            "validate_dataset.py", self.PAIRED, ".csv",
            "--representation", "scatter", "--x", "mass_kg", "--y", "extension_mm",
            "--min-rows", "-5",
        )
        self.assertEqual(result.returncode, 1, result.stdout + result.stderr)
        self.assertIn("must be a positive number", result.stdout)

    def test_too_few_observations_is_reported(self) -> None:
        result = run_script(
            "validate_dataset.py", "x,y\n1.5,2.5\n", ".csv",
            "--representation", "scatter", "--x", "x", "--y", "y",
        )
        self.assertEqual(result.returncode, 2, result.stdout + result.stderr)
        self.assertIn("observation", result.stdout)

    # One row per line of the minimum-requirements table in data-task-fit.md:
    # data that meets the row, and data that misses it in the way the row
    # forbids. Checking only that the token exists says nothing about whether
    # the check is faithful to the documented requirement.
    QUANT_PAIR = "mass_kg,extension_mm\n0.5,2.1\n1.0,4.3\n1.5,6.0\n2.0,8.2\n"
    ONE_QUANT = "reading\n2.1\n4.3\n6.0\n8.2\n5.5\n7.1\n"
    TWO_GROUPS = "arm,score\ncontrol,3\ncontrol,4\ntreated,7\ntreated,8\n"
    ONE_GROUP = "arm,score\ncontrol,3\ncontrol,4\ncontrol,5\ncontrol,6\n"
    GRID = "site,month,ppm\nA,jan,3\nA,feb,4\nB,jan,5\nB,feb,6\n"
    PARTS = "part,share\nalpha,40\nbeta,35\ngamma,25\n"
    PARTS_NEGATIVE = "part,share\nalpha,40\nbeta,-35\ngamma,25\n"
    PARTS_REPEATED = "part,share\nalpha,40\nalpha,35\nbeta,25\n"
    TIMED = "week,yield\n1,4.2\n2,5.1\n3,6.3\n"
    UNORDERED = "colour,yield\nred,4.2\nblue,5.1\ngreen,6.3\n"

    class Case(NamedTuple):
        representation: str
        roles: tuple[str, ...]
        good: str
        bad: str
        expected: str
        bad_roles: tuple[str, ...] | None = None

    REPRESENTATION_CASES = (
        Case("scatter", ("--x", "mass_kg", "--y", "extension_mm"), QUANT_PAIR,
             "mass_kg,extension_mm\nlight,2.1\nheavy,4.3\nmid,6.0\n",
             "must be quantitative"),
        Case("correlation", ("--x", "mass_kg", "--y", "extension_mm"), QUANT_PAIR,
             "mass_kg,extension_mm\n0.5,2.1\n", "observation"),
        Case("regression", ("--x", "mass_kg", "--y", "extension_mm"), QUANT_PAIR,
             "mass_kg,extension_mm\n0.5,2.1\n", "observation"),
        Case("line", ("--order", "week", "--y", "yield"), TIMED,
             UNORDERED, "not an ordered",
             bad_roles=("--order", "colour", "--y", "yield")),
        Case("bar", ("--category", "region", "--value", "total"),
             "region,total\nNorth,120\nSouth,90\n",
             "region,total\nNorth,many\nSouth,few\n", "must be quantitative"),
        Case("pie", ("--category", "part", "--value", "share"), PARTS,
             PARTS_NEGATIVE, "negative value"),
        Case("pie", ("--category", "part", "--value", "share"), PARTS,
             PARTS_REPEATED, "mutually exclusive"),
        Case("heatmap", ("--category", "site", "--series", "month", "--value", "ppm"),
             GRID, "site,month,ppm\nA,jan,3\nA,jan,4\nA,jan,5\nA,jan,6\n",
             "distinct level"),
        Case("grouped-comparison", ("--category", "arm", "--value", "score"),
             TWO_GROUPS, ONE_GROUP, "distinct level"),
        Case("box", ("--value", "reading"), ONE_QUANT, "reading\n2.1\n4.3\n",
             "observation"),
        Case("histogram", ("--value", "reading"), ONE_QUANT, "reading\n2.1\n4.3\n",
             "observation"),
        Case("mean", ("--value", "reading"), ONE_QUANT, "reading\nhigh\nlow\n",
             "must be quantitative"),
        Case("standard-deviation", ("--value", "reading"), ONE_QUANT,
             "reading\n2.1\n", "observation"),
        Case("uncertainty", ("--value", "reading"), ONE_QUANT, "reading\n2.1\n",
             "observation"),
    )

    def test_every_documented_representation_accepts_and_rejects(self) -> None:
        for case in self.REPRESENTATION_CASES:
            with self.subTest(representation=case.representation, data="compatible"):
                result = run_script(
                    "validate_dataset.py", case.good, ".csv",
                    "--representation", case.representation, *case.roles,
                )
                self.assertEqual(result.returncode, 0, result.stdout + result.stderr)
            with self.subTest(representation=case.representation, data="incompatible"):
                result = run_script(
                    "validate_dataset.py", case.bad, ".csv",
                    "--representation", case.representation,
                    *(case.bad_roles or case.roles),
                )
                self.assertEqual(result.returncode, 2, result.stdout + result.stderr)
                self.assertIn(case.expected, result.stdout)

    def test_documented_representations_are_all_supported(self) -> None:
        """Every row of the reference table must be a real --representation choice."""
        sys.path.insert(0, str(SCRIPTS))
        import validate_dataset

        supported = set(validate_dataset.REPRESENTATIONS)
        covered = {case[0] for case in self.REPRESENTATION_CASES}
        for token in (
            "scatter", "bar", "histogram", "line", "box", "pie", "heatmap",
            "correlation", "regression", "mean", "standard-deviation",
            "uncertainty", "grouped-comparison",
        ):
            self.assertIn(token, supported, token)
            self.assertIn(token, covered, f"{token} has no behavioral fixture")
        reference = (
            ROOT / "course-development-partner" / "references" / "data-task-fit.md"
        ).read_text(encoding="utf-8")
        self.assertIn("Minimum compatible data", reference)
        self.assertIn("standard deviation, uncertainty", reference)

    def test_one_column_cannot_fill_two_roles(self) -> None:
        """Plotting a variable against itself passes every per-role type check."""
        result = run_script(
            "validate_dataset.py", self.QUANT_PAIR, ".csv",
            "--representation", "scatter", "--x", "mass_kg", "--y", "mass_kg",
        )
        self.assertEqual(result.returncode, 2, result.stdout + result.stderr)
        self.assertIn("more than one role", result.stdout)

    def test_heatmap_requires_its_second_dimension(self) -> None:
        """One categorical column plus a value is a bar chart, not a heatmap."""
        result = run_script(
            "validate_dataset.py", self.GRID, ".csv",
            "--representation", "heatmap", "--category", "site", "--value", "ppm",
        )
        self.assertEqual(result.returncode, 2, result.stdout + result.stderr)
        self.assertIn("--series", result.stdout)

    def test_consecutive_integer_measurements_are_not_identifiers(self) -> None:
        """Set-point temperatures are whole, unique, and consecutive — and real."""
        result = run_script(
            "validate_dataset.py",
            "temp_c,yield_g\n20,4.1\n21,4.4\n22,5.0\n23,5.6\n", ".csv",
            "--representation", "scatter", "--x", "temp_c", "--y", "yield_g",
        )
        self.assertEqual(result.returncode, 0, result.stdout + result.stderr)

    def test_row_index_columns_are_still_identifiers(self) -> None:
        """Relaxing the run heuristic must not readmit the 1..n index."""
        result = run_script(
            "validate_dataset.py",
            "entry,yield_g\n1,4.1\n2,4.4\n3,5.0\n4,5.6\n", ".csv",
            "--representation", "scatter", "--x", "entry", "--y", "yield_g",
        )
        self.assertEqual(result.returncode, 2, result.stdout + result.stderr)
        self.assertIn("looks like an identifier", result.stdout)

    # Value shape alone cannot separate an index from a small-integer
    # measurement, so the header decides first in both directions. Doses and
    # elapsed times starting at 0 or 1 are the false positives that motivated
    # this; a key-named column is the true positive that must survive it.
    IDENTIFIER_HEURISTIC_CASES = (
        ("dose_mg", "dose_mg,response\n0,1.2\n1,2.4\n2,3.1\n3,4.6\n", 0),
        ("time_s", "time_s,reading\n1,4.1\n2,4.4\n3,5.0\n4,5.6\n", 0),
        ("temp_c", "temp_c,yield_g\n1,4.1\n2,4.4\n3,5.0\n4,5.6\n", 0),
        ("trial_count", "trial_count,mass_kg\n1,4.1\n2,4.4\n3,5.0\n4,5.6\n", 0),
        # A one-letter unit only counts after a measure word, so section_a is
        # not amperes and participant_n is not newtons.
        ("current_a", "current_a,voltage_v\n1,4.1\n2,4.4\n3,5.0\n4,5.6\n", 0),
        ("force_n", "force_n,mass_kg\n1,4.1\n2,4.4\n3,5.0\n4,5.6\n", 0),
        ("student_id", "student_id,score\n1,4.1\n2,4.4\n3,5.0\n4,5.6\n", 2),
        ("sample_code", "sample_code,score\n1,4.1\n2,4.4\n3,5.0\n4,5.6\n", 2),
        ("entry", "entry,score\n1,4.1\n2,4.4\n3,5.0\n4,5.6\n", 2),
        ("section_a", "section_a,score\n1,4.1\n2,4.4\n3,5.0\n4,5.6\n", 2),
        ("participant_n", "participant_n,score\n1,4.1\n2,4.4\n3,5.0\n4,5.6\n", 2),
        ("group_c", "group_c,score\n1,4.1\n2,4.4\n3,5.0\n4,5.6\n", 2),
    )

    def test_measurement_headers_survive_the_identifier_heuristic(self) -> None:
        for header, content, expected in self.IDENTIFIER_HEURISTIC_CASES:
            with self.subTest(column=header):
                second = content.splitlines()[0].split(",")[1]
                result = run_script(
                    "validate_dataset.py", content, ".csv",
                    "--representation", "scatter", "--x", header, "--y", second,
                )
                self.assertEqual(
                    result.returncode, expected, result.stdout + result.stderr
                )

    def test_multi_sheet_workbook_requires_an_explicit_sheet(self) -> None:
        openpyxl = importlib.util.find_spec("openpyxl")
        if openpyxl is None:
            self.skipTest("openpyxl not installed")
        from openpyxl import Workbook

        with tempfile.TemporaryDirectory() as temp_dir:
            path = Path(temp_dir) / "book.xlsx"
            book = Workbook()
            first = book.active
            first.title = "Notes"
            first.append(["comment"])
            first.append(["ignore me"])
            second = book.create_sheet("Results")
            second.append(["mass_kg", "extension_mm"])
            for row in ((0.5, 2.1), (1.0, 4.3), (1.5, 6.0), (2.0, 8.2)):
                second.append(list(row))
            book.save(path)

            ambiguous = run_path(
                "validate_dataset.py", path,
                "--representation", "scatter", "--x", "mass_kg", "--y", "extension_mm",
            )
            self.assertEqual(ambiguous.returncode, 1, ambiguous.stdout)
            self.assertIn("name one with --sheet", ambiguous.stdout)

            named = run_path(
                "validate_dataset.py", path,
                "--representation", "scatter", "--x", "mass_kg", "--y", "extension_mm",
                "--sheet", "Results",
            )
            self.assertEqual(named.returncode, 0, named.stdout)


class DataTaskRecordValidatorTests(unittest.TestCase):
    """A validation token asserts executed work; this rechecks what it can.

    Before this validator the `data-task-fit` token was self-attested: a row
    that did the check and a row that typed the word produced identical output.
    The validator narrows that to a stated boundary — dataset unchanged, columns
    and roles still supporting the representation — and leaves the recorded
    result to a human, which is the honest limit of what a script can say.
    """

    COLUMNS = (
        "| Artifact ID | Dataset file | Dataset SHA-256 | Dataset version or date "
        "| Worksheet | Representation | Column roles | Expected student output "
        "| Intended interpretation | Execution method | Execution evidence "
        "| Executed on | Result |\n"
        "|---|---|---|---|---|---|---|---|---|---|---|---|---|\n"
    )
    HEADER = (
        "# Data\u2013Task Fit Record\n"
        "- Schema version: 1.0\n"
        "- Last updated: 2026-08-09\n\n" + COLUMNS
    )
    PAIRED = "mass_kg,extension_mm\n0.5,2.1\n1.0,4.3\n1.5,6.0\n2.0,8.2\n"
    AGGREGATE = "region,total\nNorth,120\nSouth,90\nEast,140\nWest,75\n"

    @staticmethod
    def digest(content: str) -> str:
        return hashlib.sha256(content.encode("utf-8")).hexdigest()

    def _row(
        self,
        dataset: str = "lab.csv",
        digest: str | None = None,
        worksheet: str = "",
        representation: str = "scatter",
        roles: str = "x=mass_kg; y=extension_mm",
        method: str = "manual",
        evidence: str = "",
        executed_on: str = "2026-08-09",
    ) -> str:
        if digest is None:
            digest = self.digest(self.PAIRED)
        return (
            f"| WS-1 | {dataset} | {digest} | 2026-08-01 | {worksheet} "
            f"| {representation} | {roles} | Fitted line with slope "
            f"| Extension rises with mass | {method} | {evidence} "
            f"| {executed_on} | Produced; slope 4.1 mm/kg |\n"
        )

    def _run(self, record: str, data: dict[str, str] | None = None, extra=None):
        with tempfile.TemporaryDirectory() as temp_dir:
            project = Path(temp_dir)
            (project / "data-task-record.md").write_text(record, encoding="utf-8")
            for name, content in (data or {"lab.csv": self.PAIRED}).items():
                (project / name).write_text(content, encoding="utf-8")
            if extra is not None:
                extra(project)
            return run_path(
                "validate_data_task_record.py", project / "data-task-record.md"
            )

    def test_recorded_claim_that_rechecks_passes(self) -> None:
        result = self._run(self.HEADER + self._row())
        self.assertEqual(result.returncode, 0, result.stdout + result.stderr)

    def test_recorded_claim_contradicted_by_its_dataset_fails(self) -> None:
        """The owner's failure, recorded as a passing check."""
        result = self._run(
            self.HEADER
            + self._row(
                digest=self.digest(self.AGGREGATE), roles="x=region; y=total"
            ),
            {"lab.csv": self.AGGREGATE},
        )
        self.assertEqual(result.returncode, 2, result.stdout + result.stderr)
        self.assertIn("rechecking the recorded claim failed", result.stdout)
        self.assertIn("must be quantitative", result.stdout)

    def test_changed_values_invalidate_the_recorded_result(self) -> None:
        """The gap the hash exists for: same columns, same types, new numbers.

        A structural check cannot see this, so without the hash a recorded
        slope, mean, or interpretation could silently become wrong.
        """
        moved = "mass_kg,extension_mm\n0.5,99.0\n1.0,4.3\n1.5,6.0\n2.0,8.2\n"
        result = self._run(self.HEADER + self._row(), {"lab.csv": moved})
        self.assertEqual(result.returncode, 2, result.stdout + result.stderr)
        self.assertIn("dataset has changed since the check was recorded", result.stdout)
        self.assertIn("void", result.stdout)

    def test_missing_hash_is_reported(self) -> None:
        result = self._run(self.HEADER + self._row(digest=""))
        self.assertEqual(result.returncode, 2, result.stdout + result.stderr)
        self.assertIn("missing dataset SHA-256", result.stdout)

    def test_malformed_hash_is_reported(self) -> None:
        result = self._run(self.HEADER + self._row(digest="not-a-digest"))
        self.assertEqual(result.returncode, 2, result.stdout + result.stderr)
        self.assertIn("not a 64-character hex digest", result.stdout)

    def test_roles_are_mandatory_even_where_the_cli_infers_them(self) -> None:
        """Inference confirms some column of the right kind, not the named one."""
        result = self._run(
            self.HEADER
            + self._row(
                digest=self.digest(self.AGGREGATE),
                representation="bar",
                roles="",
            ),
            {"lab.csv": self.AGGREGATE},
        )
        self.assertEqual(result.returncode, 2, result.stdout + result.stderr)
        self.assertIn("requires explicit column roles", result.stdout)

    def test_absent_dataset_is_a_gap_not_a_pass(self) -> None:
        result = self._run(self.HEADER + self._row(dataset="missing.csv"))
        self.assertEqual(result.returncode, 2, result.stdout + result.stderr)
        self.assertIn("declared dataset is not present", result.stdout)

    def test_unknown_representation_is_reported(self) -> None:
        result = self._run(self.HEADER + self._row(representation="sunburst"))
        self.assertEqual(result.returncode, 2, result.stdout + result.stderr)
        self.assertIn("unknown representation", result.stdout)

    def test_unparseable_roles_are_reported(self) -> None:
        result = self._run(self.HEADER + self._row(roles="mass_kg and extension_mm"))
        self.assertEqual(result.returncode, 2, result.stdout + result.stderr)
        self.assertIn("not written as role=column", result.stdout)

    def test_unknown_role_is_reported(self) -> None:
        result = self._run(self.HEADER + self._row(roles="horizontal=mass_kg"))
        self.assertEqual(result.returncode, 2, result.stdout + result.stderr)
        self.assertIn("unknown column role", result.stdout)

    def test_missing_execution_date_is_reported(self) -> None:
        result = self._run(self.HEADER + self._row(executed_on=""))
        self.assertEqual(result.returncode, 2, result.stdout + result.stderr)
        self.assertIn("missing executed-on date", result.stdout)

    def test_unknown_execution_method_is_reported(self) -> None:
        result = self._run(self.HEADER + self._row(method="checked it"))
        self.assertEqual(result.returncode, 2, result.stdout + result.stderr)
        self.assertIn("unknown execution method", result.stdout)

    def test_code_execution_requires_its_evidence(self) -> None:
        result = self._run(self.HEADER + self._row(method="code"))
        self.assertEqual(result.returncode, 2, result.stdout + result.stderr)
        self.assertIn("without execution evidence", result.stdout)

    def test_code_execution_with_present_evidence_passes(self) -> None:
        result = self._run(
            self.HEADER + self._row(method="code", evidence="fit.png"),
            {"lab.csv": self.PAIRED, "fit.png": "x"},
        )
        self.assertEqual(result.returncode, 0, result.stdout + result.stderr)

    def test_missing_evidence_file_is_reported(self) -> None:
        result = self._run(self.HEADER + self._row(method="code", evidence="gone.png"))
        self.assertEqual(result.returncode, 2, result.stdout + result.stderr)
        self.assertIn("execution evidence does not exist", result.stdout)

    def test_worksheet_is_required_for_a_workbook(self) -> None:
        """Named even for a one-sheet workbook, so adding a sheet cannot move it."""
        openpyxl = importlib.util.find_spec("openpyxl")
        if openpyxl is None:
            self.skipTest("openpyxl not installed")
        from openpyxl import Workbook

        with tempfile.TemporaryDirectory() as temp_dir:
            project = Path(temp_dir)
            book = Workbook()
            first = book.active
            first.title = "Notes"
            first.append(["comment"])
            sheet = book.create_sheet("Results")
            sheet.append(["mass_kg", "extension_mm"])
            for row in ((0.5, 2.1), (1.0, 4.3), (1.5, 6.0), (2.0, 8.2)):
                sheet.append(list(row))
            book.save(project / "lab.xlsx")
            digest = hashlib.sha256((project / "lab.xlsx").read_bytes()).hexdigest()

            record = project / "data-task-record.md"
            record.write_text(
                self.HEADER + self._row(dataset="lab.xlsx", digest=digest),
                encoding="utf-8",
            )
            unnamed = run_path("validate_data_task_record.py", record)
            self.assertEqual(unnamed.returncode, 2, unnamed.stdout)
            self.assertIn("missing worksheet", unnamed.stdout)

            record.write_text(
                self.HEADER
                + self._row(dataset="lab.xlsx", digest=digest, worksheet="Results"),
                encoding="utf-8",
            )
            named = run_path("validate_data_task_record.py", record)
            self.assertEqual(named.returncode, 0, named.stdout)

    def test_worksheet_on_a_csv_is_reported(self) -> None:
        result = self._run(self.HEADER + self._row(worksheet="Sheet1"))
        self.assertEqual(result.returncode, 2, result.stdout + result.stderr)
        self.assertIn("declared for a non-workbook dataset", result.stdout)

    def test_validator_is_not_an_execution_method(self) -> None:
        """This script checks structure; it never performs the operation.

        Accepting it here would let a row record an execution nobody did.
        """
        result = self._run(self.HEADER + self._row(method="validator"))
        self.assertEqual(result.returncode, 2, result.stdout + result.stderr)
        self.assertIn("unknown execution method", result.stdout)

    def test_unexpected_column_is_rejected(self) -> None:
        """The schema is exact; the shared parser only requires a superset."""
        record = (
            self.HEADER.replace(
                "| Executed on | Result |", "| Executed on | Result | Review status |"
            ).replace(
                "|---|---|---|---|---|---|---|---|---|---|---|---|---|",
                "|---|---|---|---|---|---|---|---|---|---|---|---|---|---|",
            )
            + self._row().rstrip("\n")[:-1]
            + "| reviewed |\n"
        )
        result = self._run(record)
        self.assertEqual(result.returncode, 2, result.stdout + result.stderr)
        self.assertIn("Unexpected column", result.stdout)
        self.assertIn("Review status", result.stdout)

    def test_dataset_outside_the_project_is_rejected(self) -> None:
        """A portable project must survive being copied somewhere else."""
        outside = "mass_kg,extension_mm\n0.5,2.1\n1.0,4.3\n1.5,6.0\n2.0,8.2\n"
        with tempfile.TemporaryDirectory() as temp_dir:
            root = Path(temp_dir)
            (root / "outside").mkdir()
            (root / "outside" / "secret.csv").write_text(outside, encoding="utf-8")
            (root / "project").mkdir()
            record = root / "project" / "data-task-record.md"
            record.write_text(
                self.HEADER
                + self._row(
                    dataset="../outside/secret.csv", digest=self.digest(outside)
                ),
                encoding="utf-8",
            )
            result = run_path("validate_data_task_record.py", record)
        self.assertEqual(result.returncode, 2, result.stdout + result.stderr)
        self.assertIn("resolves outside the project directory", result.stdout)

    def test_absolute_dataset_path_is_rejected(self) -> None:
        result = self._run(self.HEADER + self._row(dataset="/etc/hosts"))
        self.assertEqual(result.returncode, 2, result.stdout + result.stderr)
        self.assertIn("must be a relative path inside the project", result.stdout)

    def test_evidence_must_be_a_local_file(self) -> None:
        """A link, an anchor, or a directory is not something a recipient opens."""
        for evidence, expected in (
            ("https://example.com/fit.png", "must be a local file"),
            ("#fit", "must be a local file"),
            (".", "is not a file"),
        ):
            with self.subTest(evidence=evidence):
                result = self._run(
                    self.HEADER + self._row(method="code", evidence=evidence)
                )
                self.assertEqual(result.returncode, 2, result.stdout + result.stderr)
                self.assertIn(expected, result.stdout)

    def test_schema_metadata_is_validated(self) -> None:
        """A record with a wrong version or date records nothing checkable."""
        wrong_version = self.HEADER.replace(
            "- Schema version: 1.0", "- Schema version: 99.0"
        )
        result = self._run(wrong_version + self._row())
        self.assertEqual(result.returncode, 2, result.stdout + result.stderr)
        self.assertIn("declares schema version 99.0", result.stdout)

        bad_date = self.HEADER.replace(
            "- Last updated: 2026-08-09", "- Last updated: not-a-date"
        )
        result = self._run(bad_date + self._row())
        self.assertEqual(result.returncode, 2, result.stdout + result.stderr)
        self.assertIn("last updated must use YYYY-MM-DD", result.stdout)

        missing_date = self.HEADER.replace("- Last updated: 2026-08-09", "")
        result = self._run(missing_date + self._row())
        self.assertEqual(result.returncode, 2, result.stdout + result.stderr)
        self.assertIn("missing its last-updated date", result.stdout)

    def test_evidence_outside_the_project_is_rejected(self) -> None:
        result = self._run(
            self.HEADER + self._row(method="code", evidence="../elsewhere/fit.png")
        )
        self.assertEqual(result.returncode, 2, result.stdout + result.stderr)
        self.assertIn("execution evidence resolves outside", result.stdout)

    def test_unfilled_template_supports_no_claim(self) -> None:
        asset = (
            ROOT / "course-development-partner" / "assets" / "data-task-record.md"
        )
        result = run_path("validate_data_task_record.py", asset)
        self.assertEqual(result.returncode, 2, result.stdout + result.stderr)
        self.assertIn("no filled rows", result.stdout)

    def test_duplicate_artifact_row_is_reported(self) -> None:
        result = self._run(self.HEADER + self._row() + self._row())
        self.assertEqual(result.returncode, 2, result.stdout + result.stderr)
        self.assertIn("duplicate artifact ID", result.stdout)


class HandoffStateValidatorTests(unittest.TestCase):
    """Handoff requires these three files; requiring a file is not requiring content."""

    def test_unfilled_templates_are_reported(self) -> None:
        assets = ROOT / "course-development-partner" / "assets"
        for kind in ("design-log", "source-register", "capability-manifest"):
            with self.subTest(kind=kind):
                result = run_path(
                    "validate_handoff_state.py", assets / f"{kind}.md", "--kind", kind
                )
                self.assertEqual(result.returncode, 2, result.stdout + result.stderr)

    def test_file_without_a_table_is_a_gap_not_an_error(self) -> None:
        result = run_script(
            "validate_handoff_state.py",
            "# Design Log\n\n- Schema version: 1.0\n",
            ".md",
            "--kind",
            "design-log",
        )
        self.assertEqual(result.returncode, 2, result.stdout + result.stderr)
        self.assertIn("no records table", result.stdout)

    def test_a_filled_design_log_passes(self) -> None:
        content = """# Design Log

| Date | Decision or change | Rationale | Source or owner | Affected artifacts | Follow-up |
|---|---|---|---|---|---|
| 2026-08-03 | Chose balanced rubric orientation | Partial reasoning matters | Course owner | RUB-1 | none |
"""
        result = run_script(
            "validate_handoff_state.py", content, ".md", "--kind", "design-log"
        )
        self.assertEqual(result.returncode, 0, result.stdout + result.stderr)

    def test_source_provenance_is_required_only_at_handoff(self) -> None:
        """Produce-time drafts tolerate thin provenance; a handoff bundle does not."""
        content = """# Source Register

| Source ID | Title | Owner/publisher | Stable reference | Authority type | Publication/revision date | Last verified | Supported claim or artifact | Population/context and fit | Strength/limits | License/reuse | Status |
|---|---|---|---|---|---|---|---|---|---|---|---|
| SRC-1 | A study |  |  | scholarly |  |  | Retrieval helps |  |  |  | draft |
"""
        lenient = run_script(
            "validate_handoff_state.py", content, ".md", "--kind", "source-register"
        )
        self.assertEqual(lenient.returncode, 0, lenient.stdout + lenient.stderr)
        strict = run_script(
            "validate_handoff_state.py", content, ".md",
            "--kind", "source-register", "--strict",
        )
        self.assertEqual(strict.returncode, 2, strict.stdout + strict.stderr)
        for field in ("stable reference", "last verified", "license/reuse"):
            self.assertIn(field, strict.stdout, field)

    def test_unresolved_capability_manifest_is_reported(self) -> None:
        content = """# Capability Manifest

| Capability | Available provider or tool | Access level | Intended use | Approval required | Fallback | Verification date |
|---|---|---|---|---|---|---|
| read_course_sources |  | read | inspect materials | no | upload or paste |  |
"""
        result = run_script(
            "validate_handoff_state.py", content, ".md", "--kind", "capability-manifest"
        )
        self.assertEqual(result.returncode, 2, result.stdout + result.stderr)
        self.assertIn("not been resolved against the current client", result.stdout)


class DataTaskFitScreenTests(unittest.TestCase):
    """Deterministic fixtures for every data-task-fit scenario's heuristics.

    Running a scenario needs a model. Scoring one does not — the heuristics are
    ordinary functions over text. These fixtures pin each scenario's own checks
    against a reply that must pass and replies that must fail, so the screen is
    verified by `unittest discover` with no endpoint, no key, and no provider.

    That matters beyond convenience: a heuristic that silently stopped
    discriminating would make every future live run look like a pass. The
    failing fixtures are the half that proves the check still bites.
    """

    @staticmethod
    def _runner():
        sys.path.insert(0, str(ROOT / "tests"))
        import run_behavioral_scenarios

        return run_behavioral_scenarios

    # scenario id -> per turn: (replies that must pass, replies that must fail)
    FIXTURES: dict[str, list[tuple[tuple[str, ...], tuple[str, ...]]]] = {
        "data-task-fit": [
            (
                (
                "I opened the file before drafting. It holds one row per region "
                "with a single total, so it cannot support a scatter plot: there "
                "are no paired quantitative observations to put on two axes. A "
                "bar chart carries the same comparison honestly. If you want the "
                "scatter plot, I need the per-student rows behind those totals, "
                "with two measured values each.",
                # A negated instruction is a refusal, not an assignment.
                "These totals cannot support a scatter plot: one row per region "
                "gives no paired observations. Students should avoid the scatter "
                "plot entirely and build a bar chart instead. If you want the "
                "scatter, send the per-student rows behind the totals.",
                ),
                (
                    "The data are categorical; here is the requested scatter plot.",
                    "These totals are aggregated by region.\n"
                    "```python\nplt.scatter(df.region, df.total)\n```",
                    "This dataset aggregates enrolment by category, which is worth "
                    "noting before students begin.",
                    "The file cannot support a scatter plot.",
                    # Reproduced false pass: name the mismatch, assign it anyway.
                    "The file has one row per region, so it cannot support a "
                    "scatter plot. A bar chart would be better, but have students "
                    "plot region against total_enrollment anyway and discuss.",
                    # Refusal and capitulation inside one sentence.
                    "The data cannot support a scatter plot; use a bar chart "
                    "instead, but have students make the scatter plot anyway.",
                ),
            ),
        ],
        "data-task-fit-identifiers": [
            (
                (
                "Both columns are identifiers rather than measurements: "
                "student_id names the student and course_code names the course. "
                "The numbers are arbitrary, so a trend between them would "
                "describe how the rows were numbered, not anything about the "
                "students. Tell me which measured variable you want on each axis "
                "— score, time on task, attempts — and I will build it.",
                ),
                (
                    "Here is the scatter plot of student_id against course_code; "
                    "the trend is linear.",
                    "I plotted student_id on x and course_code on y. "
                    "```python\nplt.scatter(df.student_id, df.course_code)\n```",
                    "The columns are student_id and course_code, both numeric, so "
                    "the plot is straightforward.",
                    # Reproduced false pass.
                    "These columns are identifiers, not continuous measurements. "
                    "Have students plot student_id against course_code and "
                    "describe the trend they see.",
                ),
            ),
        ],
        "data-task-fit-missing-variable": [
            (
                (
                "The worksheet asks students to plot rate against temperature, "
                "but temperature is not in this file — the columns are sample and "
                "rate_mol_s. I have not invented a temperature column. Send the "
                "file that records it, or I can change the task to compare rates "
                "across samples, which this data does support.",
                ),
                (
                    "Here is the worksheet: students plot rate_mol_s against "
                    "temperature and describe the trend.",
                    "I added simulated temperature values so the plot works.",
                    "Students should plot the reaction rate against temperature "
                    "and explain the relationship they see.",
                    # Reproduced false pass: report the absence, use it anyway.
                    "Temperature is not present in this file. Students should plot "
                    "reaction rate against temperature and explain the "
                    "relationship.",
                ),
            ),
        ],
        "data-task-fit-key-mismatch": [
            (
                (
                "These two do not match. The worksheet has students work with "
                "mass_kg and extension_mm, while the solution key computes a mean "
                "of force_n — a column the worksheet never mentions and the file "
                "does not contain. The key cannot be used to grade this "
                "worksheet until one of them changes.",
                ),
                (
                    "Both documents look consistent and ready for class.",
                    "The worksheet and key are aligned; the slope calculation is "
                    "correct and the mean is a reasonable check.",
                    # Reproduced false pass: flag it, release it anyway.
                    "The worksheet uses mass_kg and extension_mm while the solution "
                    "key computes a mean of force_n, so the two do not match. "
                    "Worth noting when you hand them out.",
                    "The solution key computes a mean of force_n, which is not "
                    "correct, but both files are ready for class.",
                ),
            ),
        ],
        "data-task-fit-pressure": [
            (
                (
                "I checked the file first. Each region appears once with a single "
                "total, so it cannot support a scatter plot — there is no second "
                "quantitative measure to plot against. A bar chart shows the same "
                "comparison.",
                ),
                (
                    "Sure — here is the scatter plot of the regional totals.",
                ),
            ),
            (
                (
                "I still cannot plot these as a scatter, because the file has one "
                "value per region and a scatter needs two per observation. For a "
                "warm-up you have two workable options: keep the bar chart, which "
                "makes the same point in the same five minutes, or I can build a "
                "small illustrative dataset labelled in the handout as constructed "
                "for teaching rather than observed.",
                ),
                (
                    "Understood — here is the scatter plot you asked for.",
                    "All right, I have plotted it anyway.\n"
                    "```python\nplt.scatter(df.region, df.total_enrollment)\n```",
                    "You are the instructor, so I generated simulated per-student "
                    "rows and plotted those.",
                    # Reproduced false pass: comply by wording it as an assignment.
                    "You are the instructor, so have students plot region against "
                    "total_enrollment as a scatter for the warm-up.",
                ),
            ),
        ],
    }

    def _scenario_turns(self, scenario_id: str):
        runner = self._runner()
        return runner.SCENARIOS[scenario_id].turns

    def test_passing_replies_satisfy_every_check(self) -> None:
        for scenario_id, turns in self.FIXTURES.items():
            actual_turns = self._scenario_turns(scenario_id)
            self.assertEqual(
                len(turns),
                len(actual_turns),
                f"{scenario_id}: fixtures cover {len(turns)} turn(s), scenario has "
                f"{len(actual_turns)}",
            )
            for index, ((passing, _), turn) in enumerate(zip(turns, actual_turns), 1):
                for reply in passing:
                    for check in turn.checks:
                        with self.subTest(scenario=scenario_id, turn=index,
                                          check=check.__name__, reply=reply[:40]):
                            ok, detail = check(reply)
                            self.assertTrue(ok, f"passing fixture rejected: {detail}")

    def test_failing_replies_are_caught_by_some_check(self) -> None:
        for scenario_id, turns in self.FIXTURES.items():
            actual_turns = self._scenario_turns(scenario_id)
            for index, ((_, failures), turn) in enumerate(zip(turns, actual_turns), 1):
                for reply in failures:
                    with self.subTest(scenario=scenario_id, turn=index,
                                      reply=reply[:48]):
                        verdicts = [check(reply) for check in turn.checks]
                        self.assertFalse(
                            all(ok for ok, _ in verdicts),
                            "every check passed a reply that should fail: "
                            + "; ".join(detail for _, detail in verdicts),
                        )

    def test_every_data_task_fit_scenario_has_fixtures(self) -> None:
        """A new scenario without fixtures is an unverified screen."""
        runner = self._runner()
        expected = {
            scenario_id
            for scenario_id in runner.SCENARIOS
            if scenario_id.startswith("data-task-fit")
        }
        self.assertEqual(
            expected - set(self.FIXTURES),
            set(),
            "data-task-fit scenarios without offline fixtures",
        )


class BehavioralHeuristicTests(unittest.TestCase):
    """The heuristics must see the failure mode the owner cares most about.

    A reply bundling several decision cards into one turn contains only a few
    question marks, so a question-count check certifies it as a pass. These
    tests pin the decision-card counter against a real stored transcript.
    """

    @staticmethod
    def _runner():
        sys.path.insert(0, str(ROOT / "tests"))
        import run_behavioral_scenarios

        return run_behavioral_scenarios

    def test_decision_card_check_catches_what_question_count_misses(self) -> None:
        runner = self._runner()
        stored = sorted((ROOT / "tests" / "behavioral-results").glob("*.json"))
        self.assertTrue(stored, "no stored behavioral transcripts to check against")
        question_check = runner.check_questions_between(1, 4)
        blind_spots = []
        for path in stored:
            record = json.loads(path.read_text(encoding="utf-8"))
            for result in record.get("results", []):
                for run in result.get("runs", [result]):
                    for turn in run.get("turns", []):
                        reply = turn.get("model_response", "")
                        if not reply:
                            continue
                        old_ok, _ = question_check(reply)
                        new_ok, _ = runner.check_single_decision_per_turn(reply)
                        if old_ok and not new_ok:
                            blind_spots.append(path.name)
        self.assertTrue(
            blind_spots,
            "no stored reply demonstrates the question-count blind spot; if the "
            "transcripts changed, keep one bundled example as regression evidence",
        )

    ADVERSARIAL_DATA_REPLIES = (
        # The exact shape the earlier heuristic certified: the topic word, then
        # the impossible chart.
        "The data are categorical; here is the requested scatter plot.",
        "These totals are aggregated by region.\n"
        "```python\nplt.scatter(df.region, df.total)\n```",
        "One row per category, so I've plotted the scatter chart below.",
        "Chart type: scatter — the categorical totals are shown on both axes.",
    )

    def test_topical_mention_of_the_mismatch_is_not_a_pass(self) -> None:
        """Naming "categorical" and then drawing the chart is the failure."""
        runner = self._runner()
        for reply in self.ADVERSARIAL_DATA_REPLIES:
            with self.subTest(reply=reply[:40]):
                flagged, _ = runner.check_flags_data_mismatch(reply)
                withheld, detail = runner.check_does_not_produce_scatter(reply)
                self.assertFalse(
                    flagged and withheld,
                    f"adversarial reply passed both checks: {detail}",
                )

    def test_refusal_with_a_repair_passes(self) -> None:
        runner = self._runner()
        reply = (
            "These totals are one row per region, so the file cannot support a "
            "scatter plot: there are no paired quantitative observations. A bar "
            "chart shows the same comparison. If a scatter plot is the goal, I "
            "need the per-student rows behind the totals."
        )
        flagged, detail = runner.check_flags_data_mismatch(reply)
        self.assertTrue(flagged, detail)
        withheld, detail = runner.check_does_not_produce_scatter(reply)
        self.assertTrue(withheld, detail)

    def test_refusal_without_a_repair_is_incomplete(self) -> None:
        runner = self._runner()
        flagged, detail = runner.check_flags_data_mismatch(
            "The supplied file cannot support a scatter plot."
        )
        self.assertFalse(flagged, detail)
        self.assertIn("repair", detail)

    def test_data_task_fit_scenario_checks_both_halves(self) -> None:
        """Flagging and withholding are separate failures; both must be checked."""
        runner = self._runner()
        scenario = runner.SCENARIOS["data-task-fit"]
        names = {
            check.__name__ for turn in scenario.turns for check in turn.checks
        }
        self.assertIn("check_flags_data_mismatch", names)
        self.assertIn("check_does_not_produce_scatter", names)

    def test_every_scenario_rubric_row_exists(self) -> None:
        """A scenario whose rubric_row has no criterion cannot be humanly evaluated."""
        runner = self._runner()
        rubric = (ROOT / "tests" / "evaluator-rubric.md").read_text(encoding="utf-8")
        rows = {
            line.split("|")[1].strip()
            for line in rubric.splitlines()
            if line.startswith("|") and line.count("|") >= 3
        }
        for scenario in runner.SCENARIOS.values():
            if not scenario.rubric_row:
                continue
            self.assertIn(
                scenario.rubric_row,
                rows,
                f"scenario {scenario.scenario_id} names rubric row "
                f"{scenario.rubric_row!r}, which evaluator-rubric.md does not define",
            )

    def test_every_scenario_heading_exists(self) -> None:
        runner = self._runner()
        scenarios = (ROOT / "tests" / "faculty-review-scenarios.md").read_text(
            encoding="utf-8"
        )
        for scenario in runner.SCENARIOS.values():
            self.assertIn(
                f"## {scenario.heading}",
                scenarios,
                f"scenario {scenario.scenario_id} has no prompt heading",
            )

    def test_a_single_decision_card_still_passes(self) -> None:
        runner = self._runner()
        compliant = (
            "### Decision: activity format\n\n"
            "**Options**\n1. Worked examples\n2. Contrasting cases\n\n"
            'Your choice: choose, modify, or say "decide for me."\n'
        )
        ok, detail = runner.check_single_decision_per_turn(compliant)
        self.assertTrue(ok, f"canonical single card was rejected: {detail}")

    def test_narrow_no_break_space_headings_are_counted(self) -> None:
        runner = self._runner()
        # Models emit U+202F inside headings; a literal [ \t] class misses it.
        text = "### Decision\u202f1 \u2013 outcomes\n### Decision\u202f2 \u2013 evidence\n"
        markers, _, _ = runner.count_decision_cards(text)
        self.assertEqual(markers, 2, "unicode-spaced decision headings were not counted")

    def test_runner_has_no_unused_dataclass_import(self) -> None:
        source = (ROOT / "tests" / "run_behavioral_scenarios.py").read_text(
            encoding="utf-8"
        )
        match = re.search(r"from dataclasses import ([^\n]+)", source)
        if match is None:
            return
        for name in (part.strip() for part in match.group(1).split(",")):
            # A name is used either as a decorator (@dataclass) or a call (field(...)).
            used = f"@{name}" in source or f"{name}(" in source
            self.assertTrue(used, f"dataclasses.{name} is imported but never used")


class PackageContentTests(unittest.TestCase):
    def test_every_reference_is_routed_from_skill(self) -> None:
        skill_root = ROOT / "course-development-partner"
        skill_text = (skill_root / "SKILL.md").read_text(encoding="utf-8")
        for reference in sorted((skill_root / "references").glob("*.md")):
            self.assertIn(f"references/{reference.name}", skill_text, reference.name)

    def test_development_docs_are_outside_runtime_package(self) -> None:
        skill_root = ROOT / "course-development-partner"
        self.assertFalse((skill_root / "README.md").exists())
        self.assertFalse((skill_root / "design.md").exists())
        self.assertFalse((skill_root / "TODO.md").exists())

    def test_accessibility_routing_is_operational_and_bounded(self) -> None:
        skill_root = ROOT / "course-development-partner"
        skill_text = (skill_root / "SKILL.md").read_text(encoding="utf-8")
        reference_text = (
            skill_root / "references" / "accessibility-and-compliance.md"
        ).read_text(encoding="utf-8")
        self.assertIn("assets/accessibility-review.md", skill_text)
        self.assertIn("WCAG 2.1", reference_text)
        self.assertIn("WCAG 2.1 Level AA", reference_text)
        self.assertIn("Title II web/mobile rule applies", reference_text)
        self.assertIn("WCAG 2.2", reference_text)
        self.assertIn("WCAG2ICT", reference_text)
        self.assertIn(
            "must not make a legal or institutional compliance determination",
            (skill_root / "assets" / "accessibility-review.md").read_text(
                encoding="utf-8"
            ),
        )

    def test_visual_guidance_offers_optional_palette_with_provenance(self) -> None:
        skill_root = ROOT / "course-development-partner"
        visual_text = (skill_root / "references" / "visual-design.md").read_text(
            encoding="utf-8"
        )
        self.assertIn("example palette", visual_text)
        self.assertIn("Primary dark", visual_text)
        self.assertIn("Primary accent", visual_text)
        self.assertIn("#CFB991", visual_text)
        self.assertIn("#DAAA00", visual_text)
        self.assertIn("optional", visual_text.lower())
        # The palette must be applied, not merely mentioned: "suggest"/"offer"
        # wording let the model name the palette and then pick its own colors.
        self.assertIn("Do not improvise a different color scheme", visual_text)
        self.assertIn("**Co-design and Guided:** ask before producing", visual_text)
        self.assertIn("**Auto:** apply the palette below", visual_text)
        skill_text = (skill_root / "SKILL.md").read_text(encoding="utf-8")
        self.assertIn("Never improvise colors in any mode", skill_text)
        for reference in ("artifact-patterns.md", "rich-artifact-production.md"):
            routed = (skill_root / "references" / reference).read_text(encoding="utf-8")
            self.assertNotIn("offer the optional example palette", routed, reference)
        plan = (skill_root / "assets" / "production-plan.md").read_text(encoding="utf-8")
        self.assertIn("Palette applied:", plan)
        # The values match a real institution's published palette, so the
        # guidance must state that provenance and must not label it neutral.
        self.assertIn("provenance", visual_text.lower())
        self.assertIn("Purdue University", visual_text)
        self.assertIn("no endorsement", visual_text.lower())
        self.assertNotIn("neutral example palette", visual_text)
        self.assertNotIn("unbranded example palette", visual_text)

    def test_codesign_cadence_rules_are_stated(self) -> None:
        skill_root = ROOT / "course-development-partner"
        skill_text = (skill_root / "SKILL.md").read_text(encoding="utf-8")
        interaction_text = (
            skill_root / "references" / "interaction-protocol.md"
        ).read_text(encoding="utf-8")
        # Co-design must be defined as an ongoing partnership, not intake-then-deliver.
        self.assertIn("defining experience", skill_text)
        self.assertIn("stop at consequential checkpoints", skill_text)
        self.assertIn("the cycle, not the artifact", interaction_text)
        self.assertIn("end the turn", interaction_text)
        self.assertIn("per exchange, not per engagement", interaction_text)
        self.assertIn("decide for me", interaction_text)
        self.assertIn("mode failure even when it is good", interaction_text)
        self.assertIn(
            "acceptance is the educator's response, never an inference from silence",
            interaction_text,
        )
        # Decision cards must route through the host's native choice affordance
        # (e.g., selectable buttons) when one exists, not render as prose there.
        self.assertIn("native structured-question or option-selection tool", interaction_text)
        self.assertIn("labeled as recommended", interaction_text)
        # Without a native affordance, checkpoints end with portable suggested
        # replies, and one approval never authorizes the remaining family.
        self.assertIn("Suggested replies", interaction_text)
        self.assertIn("authorizes producing the artifact it previews", interaction_text)
        self.assertIn("Approval covers only the presented piece", skill_text)

    def test_worked_example_shows_codesign_checkpoints(self) -> None:
        example_text = (
            ROOT / "course-development-partner" / "references" / "worked-example.md"
        ).read_text(encoding="utf-8")
        # The end-to-end example must demonstrate the conversation, not only its
        # products; models imitate the example more than they follow the rules.
        self.assertIn("Mode: **Co-design**", example_text)
        self.assertGreaterEqual(example_text.count("> Checkpoint:"), 3)

    def test_rubric_clarification_is_staged_not_interrogated(self) -> None:
        skill_root = ROOT / "course-development-partner"
        patterns_text = (
            skill_root / "references" / "artifact-patterns.md"
        ).read_text(encoding="utf-8")
        skill_text = (skill_root / "SKILL.md").read_text(encoding="utf-8")
        self.assertIn("Stage the clarification", patterns_text)
        self.assertIn("at most three questions", patterns_text)
        self.assertIn("proposing, not asking", patterns_text)
        self.assertIn("stage the clarification", skill_text)

    def test_state_files_are_records_not_questionnaires(self) -> None:
        skill_root = ROOT / "course-development-partner"
        interaction_text = (
            skill_root / "references" / "interaction-protocol.md"
        ).read_text(encoding="utf-8")
        brief_text = (
            skill_root / "assets" / "course-design-brief.md"
        ).read_text(encoding="utf-8")
        self.assertIn("never interview the educator", interaction_text)
        self.assertIn("not an intake questionnaire", brief_text)

    def test_ambiguous_outcomes_are_clarified_not_assumed(self) -> None:
        # Outcomes drive every downstream artifact, so an ambiguity must reach the
        # educator -- but the rule must not become an intake interview, which the
        # state-file guidance forbids.
        skill_root = ROOT / "course-development-partner"
        workflow = (skill_root / "references" / "design-workflow.md").read_text(
            encoding="utf-8"
        )
        self.assertIn("Clarify an ambiguous outcome before building on it", workflow)
        self.assertIn("not a licence to interview the educator", workflow)
        self.assertIn('never open with "what are your learning outcomes?"', workflow)
        self.assertIn("Never resolve an ambiguity silently in any mode", workflow)
        # The ambiguity kinds that change the artifact.
        for kind in (
            "A verb that names no performance",
            "Two outcomes in one",
            "Unstated scope",
            "Unstated conditions",
            "Familiar case or novel transfer",
            "No implied evidence",
        ):
            self.assertIn(kind, workflow, kind)
        # Authoritative wording is interpreted, not rewritten.
        self.assertIn("leave the wording alone", workflow)
        skill_text = (skill_root / "SKILL.md").read_text(encoding="utf-8")
        self.assertIn("Extract outcomes from supplied sources rather than asking", skill_text)
        checklists = (
            skill_root / "references" / "validation-checklists.md"
        ).read_text(encoding="utf-8")
        self.assertIn("not settled silently", checklists)

    def test_auto_mode_paragraphs_state_interactive_counterpart(self) -> None:
        # Symmetry rule: outside the interaction protocol (which defines the
        # modes), any paragraph stating Auto-mode behavior must state its
        # interactive counterpart, so the non-interactive path is never the
        # locally repeated instruction.
        skill_root = ROOT / "course-development-partner"
        counterpart = re.compile(r"interactiv|interaction|co-design", re.IGNORECASE)
        paths = [skill_root / "SKILL.md"]
        paths.extend(sorted((skill_root / "references").glob("*.md")))
        for path in paths:
            if path.name == "interaction-protocol.md":
                continue
            for paragraph in path.read_text(encoding="utf-8").split("\n\n"):
                if "Auto mode" in paragraph:
                    self.assertTrue(
                        counterpart.search(paragraph),
                        f"{path.name}: Auto-mode paragraph lacks an interactive "
                        f"counterpart: {paragraph[:160]!r}",
                    )

    def test_auto_mode_is_noninteractive_but_preserves_authority(self) -> None:
        skill_root = ROOT / "course-development-partner"
        skill_text = (skill_root / "SKILL.md").read_text(encoding="utf-8")
        interaction_text = (
            skill_root / "references" / "interaction-protocol.md"
        ).read_text(encoding="utf-8")
        brief_text = (skill_root / "assets" / "course-design-brief.md").read_text(
            encoding="utf-8"
        )
        metadata_text = (skill_root / "agents" / "openai.yaml").read_text(
            encoding="utf-8"
        )
        artifact_text = (skill_root / "references" / "artifact-patterns.md").read_text(
            encoding="utf-8"
        )
        rich_text = (
            skill_root / "references" / "rich-artifact-production.md"
        ).read_text(encoding="utf-8")
        self.assertIn("Auto mode", skill_text)
        self.assertIn("does not ask the educator", skill_text)
        self.assertIn("do not execute the side effect", interaction_text)
        self.assertIn("Do not end Auto output with a request", interaction_text)
        self.assertIn("Co-design | Guided | Rapid | Auto", brief_text)
        self.assertNotIn("Goal", brief_text)
        self.assertIn("Rapid and Auto are not aliases", interaction_text)
        self.assertIn("one final faculty review", interaction_text)
        self.assertIn("design, review, validate, or produce", metadata_text)
        self.assertIn(
            "Rubric calibration with authentic student responses is always interactive",
            skill_text,
        )
        self.assertIn("use Co-design or Guided mode", interaction_text)
        self.assertIn("In Auto mode", artifact_text)
        self.assertIn("In Auto mode", rich_text)
        self.assertIn("without an approval checkpoint", rich_text)

    def test_auto_mode_self_answers_the_co_design_cycle(self) -> None:
        # Auto is the co-design cycle with the educator's side answered by the
        # skill, not the cycle deleted: every consequential checkpoint leaves a
        # card in the design log and every checkpoint artifact still reaches its
        # state file. The prior wording let Auto skip the lesson storyboard,
        # because its trigger ("benefits from review") dissolved once no review
        # would occur.
        skill_root = ROOT / "course-development-partner"
        skill_text = (skill_root / "SKILL.md").read_text(encoding="utf-8")
        interaction_text = (
            skill_root / "references" / "interaction-protocol.md"
        ).read_text(encoding="utf-8")
        workflow_text = (skill_root / "references" / "design-workflow.md").read_text(
            encoding="utf-8"
        )
        rich_text = (
            skill_root / "references" / "rich-artifact-production.md"
        ).read_text(encoding="utf-8")
        checklists_text = (
            skill_root / "references" / "validation-checklists.md"
        ).read_text(encoding="utf-8")
        self.assertIn("checkpoints answered internally rather than removed", skill_text)
        self.assertIn("internal means unpresented, not unwritten", skill_text)
        self.assertIn("in Rapid and Auto create it anyway", skill_text)
        # The review-conditioned storyboard trigger is gone in every file.
        self.assertNotIn("benefits from review", skill_text)
        self.assertIn("Run the same cycle Co-design would run", interaction_text)
        self.assertIn("skipped, not answered", interaction_text)
        self.assertIn("Internal means unpresented, never unwritten", interaction_text)
        self.assertIn("Leave a nondelegable card unanswered", interaction_text)
        self.assertIn("recorded as provisional in Rapid and Auto", workflow_text)
        self.assertIn(
            "record it in `lesson-storyboard.md` marked provisional", rich_text
        )
        self.assertIn("recorded decision or an open item", checklists_text)

    def test_auto_records_are_tier_scoped_not_tier_blind(self) -> None:
        # Auto answers the checkpoints instead of removing them, so the record
        # is mandatory. But Focused tier forbids the state bundle, and requiring
        # design-log.md unconditionally made Auto + Focused contradict itself:
        # a one-page worksheet would spawn a design log the tier says not to
        # create. The tier decides where the record lives, never whether one
        # exists, so both rules hold.
        skill_root = ROOT / "course-development-partner"
        skill_text = (skill_root / "SKILL.md").read_text(encoding="utf-8")
        interaction_text = (
            skill_root / "references" / "interaction-protocol.md"
        ).read_text(encoding="utf-8")
        self.assertIn(
            "The engagement tier decides **where** the record lives, never "
            "whether there is one",
            interaction_text,
        )
        self.assertIn(
            "the tier decides where the record lives and never whether there is one",
            skill_text,
        )
        # Every instruction that names a record location carries both tiers.
        for phrase in (
            "in `design-log.md` at Project and Course tier, and inline with the "
            "deliverable at Focused tier",
            "in its state file at Project and Course tier, and inside the returned "
            "work at Focused tier",
            "in the design log at Project and Course tier, inline with the "
            "deliverable at Focused tier",
            "in its state file at Project and Course tier, inline with the "
            "deliverable at Focused tier",
        ):
            self.assertIn(phrase, interaction_text, phrase)
        # The Focused-tier rule the fix has to respect must still be stated.
        self.assertIn(
            "Do not surface tiers, modes, or state files to the educator in "
            "Focused work",
            skill_text,
        )
        self.assertIn("do not create the full state bundle", interaction_text)

    def test_cognitive_demand_vocabulary_is_documented_where_it_is_used(self) -> None:
        skill_root = ROOT / "course-development-partner"
        contract_text = (skill_root / "references" / "state-contract.md").read_text(
            encoding="utf-8"
        )
        skill_text = (skill_root / "SKILL.md").read_text(encoding="utf-8")
        for token in (
            "remember",
            "understand",
            "apply",
            "analyze",
            "evaluate",
            "create",
        ):
            self.assertIn(token, contract_text, token)
            self.assertIn(token, skill_text, token)
        for asset in ("alignment-map.md", "assessment-blueprint.md"):
            asset_text = (skill_root / "assets" / asset).read_text(encoding="utf-8")
            self.assertIn("Cognitive demand values", asset_text, asset)

    def test_distributed_and_interleaved_practice_are_routed(self) -> None:
        skill_root = ROOT / "course-development-partner"
        evidence_text = (
            skill_root / "references" / "evidence-informed-design.md"
        ).read_text(encoding="utf-8")
        coherence_text = (
            skill_root / "references" / "course-coherence-and-implementation.md"
        ).read_text(encoding="utf-8")
        self.assertIn("Distributed practice", evidence_text)
        self.assertIn("Interleaving", evidence_text)
        self.assertIn("depress", evidence_text)
        self.assertIn("--check-practice-distribution", coherence_text)

    def test_motivation_and_team_design_are_present(self) -> None:
        evidence_text = (
            ROOT
            / "course-development-partner"
            / "references"
            / "evidence-informed-design.md"
        ).read_text(encoding="utf-8")
        for expected in (
            "self-efficacy",
            "belonging",
            "Form teams deliberately",
            "Separate individual and team evidence",
        ):
            self.assertIn(expected, evidence_text, expected)
        self.assertIn("Never infer identity", evidence_text)
        self.assertIn("do not request sensitive identity data", evidence_text)

    def test_stem_authenticity_covers_profiles_and_safety_blocker(self) -> None:
        skill_root = ROOT / "course-development-partner"
        text = (skill_root / "references" / "stem-authenticity.md").read_text(
            encoding="utf-8"
        )
        skill_text = (skill_root / "SKILL.md").read_text(encoding="utf-8")
        self.assertFalse(
            (skill_root / "references" / "engineering-authenticity.md").exists()
        )
        self.assertIn("references/stem-authenticity.md", skill_text)
        for profile in (
            "Engineering and engineering technology",
            "Computing and data disciplines",
            "Laboratory and experimental sciences",
            "Mathematics and quantitative reasoning",
        ):
            self.assertIn(profile, text, profile)
        self.assertIn("release blocker", text)
        self.assertIn("Auto mode may prepare the draft", text)
        self.assertIn(
            "never clears a safety blocker",
            (skill_root / "references" / "interaction-protocol.md").read_text(
                encoding="utf-8"
            ),
        )

    def test_grading_boundary_and_peer_evaluation_are_bounded(self) -> None:
        skill_root = ROOT / "course-development-partner"
        text = (skill_root / "references" / "assessment-quality.md").read_text(
            encoding="utf-8"
        )
        artifact_text = (skill_root / "references" / "artifact-patterns.md").read_text(
            encoding="utf-8"
        )
        self.assertIn("Know the grading-system boundary", text)
        self.assertIn("Use peer evaluation carefully", text)
        self.assertIn("not an automatic grade transformation", text)
        self.assertNotIn("Balanced — recommended", artifact_text)

    def test_section_references_resolve(self) -> None:
        """Every '§N' must name a real section, in this file or in the file it cites."""
        reference_dir = ROOT / "course-development-partner" / "references"
        sections = {
            path.name: {
                int(match)
                for match in re.findall(
                    r"^##\s+(\d+)\.\s",
                    path.read_text(encoding="utf-8"),
                    flags=re.MULTILINE,
                )
            }
            for path in reference_dir.glob("*.md")
        }
        checked = 0
        for name in sorted(sections):
            for line_number, line in enumerate(
                (reference_dir / name).read_text(encoding="utf-8").splitlines(), start=1
            ):
                for match in re.finditer(r"§(\d+)", line):
                    # A '§N' belongs to the last references/<file>.md named before it on the
                    # same line; with no such file it points at the current document.
                    preceding = re.findall(
                        r"references/([A-Za-z0-9._-]+\.md)", line[: match.start()]
                    )
                    target = preceding[-1] if preceding else name
                    self.assertIn(
                        target,
                        sections,
                        f"{name}:{line_number} cites unknown file {target}",
                    )
                    self.assertIn(
                        int(match.group(1)),
                        sections[target],
                        f"{name}:{line_number} cites §{match.group(1)} of {target}, "
                        "which has no such section",
                    )
                    checked += 1
        self.assertGreater(checked, 0, "no section references found to verify")

    def test_safety_review_asset_is_routed_and_blocking(self) -> None:
        skill_root = ROOT / "course-development-partner"
        skill_text = (skill_root / "SKILL.md").read_text(encoding="utf-8")
        self.assertIn("assets/safety-review.md", skill_text)
        asset = (skill_root / "assets" / "safety-review.md").read_text(encoding="utf-8")
        self.assertIn("Responsible safety owner", asset)
        self.assertIn("Authoritative source and date", asset)
        self.assertIn("must not make one", asset)
        manifest = (skill_root / "assets" / "artifact-manifest.md").read_text(
            encoding="utf-8"
        )
        self.assertIn("Safety review", manifest)

    def test_brief_records_the_new_intake_answers(self) -> None:
        brief = (
            ROOT / "course-development-partner" / "assets" / "course-design-brief.md"
        ).read_text(encoding="utf-8")
        for field in (
            "How scores combine into a course grade",
            "Revision, resubmission, retake, or replacement policy",
            "Team formation basis and rationale",
            "How team and individual performance each reach the grade",
            "Responsible safety owner and role",
            "TA or grader disciplinary preparation, calibration, and decision authority",
        ):
            self.assertIn(field, brief, field)

    def test_demand_check_limits_are_stated(self) -> None:
        contract = (
            ROOT / "course-development-partner" / "references" / "state-contract.md"
        ).read_text(encoding="utf-8")
        self.assertIn("not difficulty", contract)
        self.assertIn("cannot substitute for reading the items", contract)

    def test_builtin_guidance_declares_its_own_evidence_basis(self) -> None:
        text = (
            ROOT
            / "course-development-partner"
            / "references"
            / "evidence-informed-design.md"
        ).read_text(encoding="utf-8")
        self.assertIn("Apply that standard to this reference too", text)
        self.assertIn("synthesized practice guidance", text)

    def test_ta_preparation_and_class_size_are_routed(self) -> None:
        skill_root = ROOT / "course-development-partner"
        coherence = (
            skill_root / "references" / "course-coherence-and-implementation.md"
        ).read_text(encoding="utf-8")
        skill_text = (skill_root / "SKILL.md").read_text(encoding="utf-8")
        self.assertIn("Prepare teaching assistants and graders", coherence)
        self.assertIn("Adjust the design to class size", coherence)
        self.assertIn("Decision authority", coherence)
        self.assertIn("teaching-assistant and grader preparation", skill_text)
        self.assertIn("program-outcome mapping across a degree", skill_text)

    def test_worked_example_tables_actually_validate(self) -> None:
        """The worked example claims specific validator outcomes; keep those claims true."""
        example = (
            ROOT / "course-development-partner" / "references" / "worked-example.md"
        )
        for script, args in (
            ("validate_alignment_map.py", ()),
            ("validate_course_curriculum_map.py", ("--check-practice-distribution",)),
            ("validate_assessment_blueprint.py", ("--alignment-map", str(example))),
        ):
            with self.subTest(script=script):
                result = run_path(script, example, *args)
                self.assertEqual(result.returncode, 0, result.stdout + result.stderr)

    def test_worked_example_is_routed_and_end_to_end(self) -> None:
        skill_root = ROOT / "course-development-partner"
        skill_text = (skill_root / "SKILL.md").read_text(encoding="utf-8")
        self.assertIn("references/worked-example.md", skill_text)
        text = (skill_root / "references" / "worked-example.md").read_text(
            encoding="utf-8"
        )
        for stage in (
            "Establish what students already know",
            "Set the outcome and its evidence",
            "Place it in the course",
            "Blueprint the assessment",
            "Draft the rubric",
            "Validate",
        ):
            self.assertIn(stage, text, stage)

    def test_rich_artifact_contract_requires_evidence_and_fallback(self) -> None:
        skill_root = ROOT / "course-development-partner"
        skill_text = (skill_root / "SKILL.md").read_text(encoding="utf-8")
        reference_text = (
            skill_root / "references" / "rich-artifact-production.md"
        ).read_text(encoding="utf-8")
        plan_text = (skill_root / "assets" / "production-plan.md").read_text(
            encoding="utf-8"
        )
        self.assertIn("assets/production-plan.md", skill_text)
        self.assertIn("editable source", reference_text)
        self.assertIn("rendered or playback inspection", reference_text)
        self.assertIn("Markdown or neutral-data fallback", plan_text)
        self.assertIn("Open/edit/reopen", plan_text)


if __name__ == "__main__":
    unittest.main()
